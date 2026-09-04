"""
sheets_manager.py
Módulo de acceso a Google Sheets para Sistema de Asistencia RRHH - Quski
"""

import gspread
from gspread.utils import rowcol_to_a1
from google.oauth2.service_account import Credentials
import pandas as pd
from datetime import datetime, date, timedelta
import hashlib
import hmac
import os
import base64
import time
from zoneinfo import ZoneInfo

# Segundos que se reutiliza una lectura de una hoja antes de volver a pedirla.
# Google Sheets limita a ~60 lecturas por minuto y por usuario; sin esta caché
# una sola pantalla puede gastar 5 o 6 lecturas y agotar la cuota.
_CACHE_TTL = 90

# Columna interna con el número de fila real en la hoja de cálculo. Empieza con
# guion bajo para distinguirla de las columnas de datos; las tablas de pantalla
# la ocultan.
FILA_HOJA = "_fila_hoja"


def fila_de_hoja(df, idx) -> int:
    """Número de fila en la hoja para una fila del DataFrame.

    Usa la columna que guardó el lector. Solo cae en el cálculo por posición
    —idx + 2— cuando el DataFrame viene de otra parte y no la trae.
    """
    try:
        if FILA_HOJA in df.columns:
            return int(df.at[idx, FILA_HOJA])
    except Exception:
        pass
    return int(idx) + 2

ZONA_POR_DEFECTO = "America/Guayaquil"

# ── Reintentos ante fallos transitorios de Google ─────────────────────────────
# La API de Sheets devuelve 503 ("service currently unavailable") o 500 de vez
# en cuando, sin que haya nada mal en la petición: son caídas momentáneas del
# lado de Google. Sin reintentos, un 503 de un segundo deja a todo el personal
# sin poder entrar al sistema.
REINTENTOS_MAX = 4
ESPERA_BASE_SEG = 1.5

_ERRORES_TRANSITORIOS = ("503", "500", "502", "504", "429",
                         "currently unavailable", "internal error",
                         "backend error", "quota exceeded", "rate_limit")


_ERRORES_CUOTA = ("429", "quota exceeded", "rate_limit", "resource_exhausted",
                  "too many requests")


def es_error_cuota(e) -> bool:
    """True si Google rechazó la llamada por exceso de consultas."""
    texto = f"{type(e).__name__}: {e}".lower()
    return any(marca in texto for marca in _ERRORES_CUOTA)


def es_error_transitorio(e) -> bool:
    texto = f"{type(e).__name__}: {e}".lower()
    return any(marca in texto for marca in _ERRORES_TRANSITORIOS)


def con_reintentos(fn, *args, **kwargs):
    """Ejecuta una llamada a la API reintentando ante fallos pasajeros.

    Espera cada vez más entre intentos (1.5s, 3s, 6s) para no empeorar una
    sobrecarga. Los errores que no son transitorios —permisos, hoja inexistente,
    petición mal formada— se propagan de inmediato: reintentarlos no sirve.
    """
    ultimo = None
    for intento in range(REINTENTOS_MAX):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            # La cuota agotada NO se reintenta: cada reintento consume otra
            # llamada del mismo minuto y hunde más el límite. Se propaga de
            # una vez para que la pantalla lo explique y la persona espere.
            if es_error_cuota(e):
                raise
            if not es_error_transitorio(e) or intento == REINTENTOS_MAX - 1:
                raise
            ultimo = e
            time.sleep(ESPERA_BASE_SEG * (2 ** intento))
    if ultimo:
        raise ultimo


def zona_horaria(config: dict | None = None) -> ZoneInfo:
    """Zona horaria configurada (America/Guayaquil por defecto)."""
    nombre = (config or {}).get("Zona_Horaria", ZONA_POR_DEFECTO) or ZONA_POR_DEFECTO
    try:
        return ZoneInfo(nombre)
    except Exception:
        return ZoneInfo(ZONA_POR_DEFECTO)


def ahora_local(config: dict | None = None) -> datetime:
    """Fecha y hora actuales en la zona horaria de la empresa.

    Importante: el servidor de Streamlit Cloud corre en UTC, así que
    datetime.now() devuelve una hora cinco horas adelantada respecto a Ecuador.
    Todo registro de asistencia debe pasar por aquí."""
    return datetime.now(zona_horaria(config))


def hoy_local(config: dict | None = None) -> date:
    """Fecha de hoy en la zona horaria de la empresa."""
    return ahora_local(config).date()

# ── Constantes ────────────────────────────────────────────────────────────────
SPREADSHEET_ID = "1gaPrP95SF0xat7xRs94CMyH22LeIoolEom1OjCDAA2I"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

HEADERS = {
    "Empleados":    ["ID_Empleado", "Nombre", "Email", "Area", "Email_Jefe", "Horario_Inicio", "Horario_Fin",
                     "Fecha_Ingreso", "Dias_Tomados_Inicial", "Modalidad", "Dia_Oficina"],
    "Asistencia":   ["Fecha", "ID_Empleado", "Nombre", "Hora_Entrada", "Hora_Salida", "Estado",
                     "Minutos_Atraso", "Observaciones", "Aviso_Salida",
                     "Dispositivo_Entrada", "Dispositivo_Salida"],
    "Permisos":     ["ID_Permiso", "Fecha", "ID_Empleado", "Horas_Solicitadas", "Motivo", "Estado",
                     "Horas_Usadas_Mes", "Aprobado_Por",
                     "Aprobado_Jefe", "Fecha_Aprob_Jefe", "Aprobado_RRHH", "Fecha_Aprob_RRHH",
                     "Motivo_Rechazo", "Rechazado_Por"],
    "Vacaciones":   ["ID_Vacacion", "ID_Empleado", "Fecha_Inicio", "Fecha_Fin", "Dias_Habiles", "Estado",
                     "Aprobado_Por",
                     "Aprobado_Jefe", "Fecha_Aprob_Jefe", "Aprobado_RRHH", "Fecha_Aprob_RRHH",
                     "Motivo_Rechazo", "Rechazado_Por", "Dias_Calendario",
                     "Motivo", "Reemplazo"],
    "Horas_Extras": ["ID", "Fecha", "ID_Empleado", "Horas_Extra", "Motivo", "Aprobado_Por", "Estado"],
    "Configuracion":["Key", "Valor"],
    "Usuarios":          ["ID_Empleado", "Password_Hash", "Rol"],
    "Llamados_Atencion": ["ID_Llamado", "Fecha", "ID_Empleado", "Nombre", "Tipo", "Motivo",
                          "Atrasos_Acumulados", "Registrado_Por", "Estado"],
    # Formulario "Conozca a su Empleado" (KYE) — base de la matriz de riesgo
    "KYE_Empleado": ["ID_Empleado", "Cedula", "Fecha_Nacimiento", "Nacionalidad",
                     "Ciudad_Nacimiento", "Estado_Civil", "Sexo", "Num_Hijos",
                     "Provincia", "Canton", "Parroquia", "Direccion",
                     "Telefono_Domicilio", "Celular", "Email_Personal",
                     "Conyuge_Nombre", "Conyuge_Cedula", "Conyuge_Trabaja",
                     "Conyuge_Relacion", "Conyuge_Empresa", "Conyuge_Actividad",
                     "Conyuge_Cargo", "Conyuge_Anios_Cargo",
                     "Es_PEP", "PEP_Cargo", "Familiar_PEP", "Familiar_PEP_Detalle",
                     "Otros_Ingresos", "Otros_Ingresos_Detalle",
                     "Doc_Hoja_Vida", "Doc_Cedula", "Doc_Cedula_Conyuge",
                     "Doc_Papeleta", "Doc_Papeleta_Conyuge", "Doc_Ref_Laborales",
                     "Doc_Ref_Personales", "Doc_Servicio_Basico",
                     "Doc_Declaracion_Patrimonial",
                     "Fecha_Formulario", "Observaciones", "Registrado_Por", "Actualizado",
                     "Aviso_Tecnologia"],
    "Score_Buro": ["ID_Score", "ID_Empleado", "Fecha", "Score", "Tipo", "Fuente",
                   "Observaciones", "Registrado_Por"],
    "Reconocimientos": ["ID_Reconocimiento", "Fecha", "ID_Empleado", "Nombre", "Tipo",
                        "Motivo", "Otorgado_Por", "Estado"],
    # Repositorio: el archivo vive en Drive, aquí queda el enlace y su estado
    "Documentos": ["ID_Documento", "Fecha", "ID_Empleado", "Nombre", "Categoria",
                   "Tipo_Documento", "Nombre_Archivo", "Drive_ID", "Drive_Link",
                   "Subido_Por", "Estado", "Revisado_Por", "Fecha_Revision",
                   "Referencia", "Observaciones"],
    "Capacitaciones": ["ID_Curso", "ID_Empleado", "Nombre", "Curso", "Institucion",
                       "Financiamiento", "Fecha_Inicio", "Fecha_Fin", "Estado",
                       "Horas", "Costo", "Documento_ID", "Registrado_Por",
                       "Observaciones"],
    "Titulos": ["ID_Titulo", "ID_Empleado", "Nombre", "Nivel", "Titulo",
                "Institucion", "Anio_Obtencion", "Registro_SENESCYT",
                "Documento_ID", "Estado_Validacion", "Validado_Por",
                "Fecha_Validacion", "Registrado_Por", "Observaciones"],
}

# Niveles de formación, de menor a mayor. El puntaje se usa para reconocer la
# formación en la matriz de riesgo: a mayor nivel, menor riesgo operativo.
NIVELES_TITULO = {
    "Bachiller": 1,
    "Técnico / Tecnólogo": 2,
    "Tercer nivel (grado)": 3,
    "Especialización": 4,
    "Cuarto nivel (maestría)": 5,
    "Doctorado (PhD)": 6,
}

CATEGORIAS_DOCUMENTO = [
    "Documento KYE",
    "Hoja de vida",
    "Título académico",
    "Certificado de curso",
    "Certificado laboral",
    "Otro",
]

TIPOS_DOC_KYE = [
    "a) Hoja de vida", "b) Copia de cédula", "c) Cédula del cónyuge",
    "d) Papeleta de votación", "e) Papeleta del cónyuge",
    "f) Referencias laborales", "g) Referencias personales",
    "h) Planilla de servicio básico", "i) Declaración patrimonial",
]

ESTADOS_CURSO = ["En curso", "Finalizado", "Abandonado", "Inscrito"]
FINANCIAMIENTO_CURSO = ["Pagado por la empresa", "Cofinanciado", "Por cuenta del empleado"]

# Tipos de reconocimiento que se registran en el expediente
TIPOS_RECONOCIMIENTO = [
    "Carta de felicitación",
    "Reconocimiento por metas",
    "Mención por atención al cliente",
    "Reconocimiento por antigüedad",
    "Otro",
]

CONFIG_DEFAULTS = {
    "Horario_Inicio":              "09:00",
    "Horario_Fin":                 "17:30",
    "Tolerancia_Minutos":          "0",
    "Horas_Permiso_Mensual":       "3",
    "Email_RRHH":                  "rrhh@quski.ec",
    "Zona_Horaria":                "America/Guayaquil",
    "Tardanzas_Llamado_Verbal":    "3",
    "Tardanzas_Llamado_Escrito":   "5",
    "Tardanzas_Suspension":        "8",
    # Vacaciones: 15 días calendario al año; desde el año indicado se suma un
    # día por cada año adicional de antigüedad, con un techo.
    "Dias_Vacaciones_Base":        "15",
    "Anio_Inicio_Dia_Adicional":   "5",
    "Max_Dias_Vacaciones":         "30",
    # Minutos de gracia antes del horario de fin sin que la salida se considere
    # anticipada. En 0, cualquier salida antes de la hora dispara el aviso.
    "Tolerancia_Salida_Minutos":   "0",
    # Horas antes del horario de inicio en que todavía se acepta marcar entrada.
    # La ventana va de (Horario_Inicio − este margen) hasta Horario_Fin.
    "Margen_Registro_Entrada_Horas": "3",
    # Feriados: fechas AAAA-MM-DD separadas por coma. Sin esto, un feriado
    # nacional aparece en el panel como si todo el equipo hubiera faltado.
    "Feriados":                    "",
    # ── Modelo de riesgo operativo por asesor ──────────────────────────────
    # Pesos de cada factor (suman 100; el sistema normaliza si no).
    "Riesgo_Peso_Buro":            "30",
    "Riesgo_Peso_PEP":             "10",
    "Riesgo_Peso_Documentos":      "15",
    "Riesgo_Peso_Familiar":        "10",
    "Riesgo_Peso_Disciplina":      "25",
    "Riesgo_Peso_Antiguedad":      "10",
    # Cortes del score de buró (mayor score = mejor comportamiento crediticio)
    "Buro_Score_Bueno":            "800",
    "Buro_Score_Malo":             "400",
    # Umbrales del nivel de riesgo sobre el puntaje 0-100
    "Riesgo_Umbral_Medio":         "30",
    "Riesgo_Umbral_Alto":          "55",
    "Riesgo_Umbral_Critico":       "75",
    # Calibración de factores
    "Riesgo_Hijos_Tope":              "4",
    "Riesgo_Puntos_Disciplina_Tope":  "12",
    "Riesgo_Peso_Reconocimiento":     "1.5",
    "Riesgo_Antiguedad_Anios":        "2",
    "Riesgo_Peso_Formacion":          "10",
    # Carpeta de Google Drive donde se guardan los documentos del personal.
    # Debe ser de una persona real y estar compartida como Editor con la
    # cuenta de servicio: los service accounts no tienen espacio propio.
    "Drive_Carpeta_ID":               "",
    # Área de sistemas: recibe los datos de contacto del personal para
    # bloquearlos en el BPM y que nadie pueda usarlos en un cliente.
    "Email_Tecnologia":               "tecnologia@quski.ec",
}

# Datos de contacto personales del empleado que sistemas debe bloquear en el
# BPM. Si un asesor pudiera registrar un cliente con su propio celular o
# dirección, tendría cómo autogestionarse una operación.
CAMPOS_BLOQUEO_BPM = [
    ("Email_Personal",     "Correo electrónico personal"),
    ("Celular",            "Número de celular"),
    ("Telefono_Domicilio", "Teléfono de domicilio"),
    ("Direccion",          "Dirección de domicilio"),
]

# ── Estados del flujo de aprobación ───────────────────────────────────────────
# Pendiente_Jefe  → esperando al jefe inmediato
# Pendiente_RRHH  → jefe aprobó, esperando a RRHH
# Aprobado        → aprobado por ambos (o solo RRHH si no hay jefe asignado)
# Rechazado       → rechazado por el jefe o por RRHH
# Modalidad de trabajo. Determina contra qué se compara la asistencia de cada
# persona: no es lo mismo que no haya marca de alguien que debe estar en planta
# que de alguien que trabaja desde su casa.
MOD_PRESENCIAL  = "Presencial"     # todos los días en la oficina o la fábrica
MOD_TELETRABAJO = "Teletrabajo"    # todos los días desde fuera
MOD_MIXTO       = "Mixto"          # teletrabajo con un día fijo de oficina
MODALIDADES = (MOD_PRESENCIAL, MOD_TELETRABAJO, MOD_MIXTO)

DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]


EST_PEND_JEFE = "Pendiente_Jefe"
EST_PEND_RRHH = "Pendiente_RRHH"
EST_APROBADO  = "Aprobado"
EST_RECHAZADO = "Rechazado"

# Tipo reservado para el registro automático de atrasos. Se guarda en la misma
# hoja de Llamados_Atencion pero NO es una sanción: queda separado de los
# llamados Verbal / Escrito / Suspensión que emite RRHH a mano.
TIPO_TARDANZA = "Tardanza (registro automático)"

# Causales operativas de llamado de atención, además de la escala disciplinaria
# clásica. Auditoría y los jefes de área registran sobre estas.
CAUSALES_LLAMADO = [
    "Errores de tasación",
    "Errores de gestión",
    "Mala atención al cliente",
    "Errores en cierre de caja",
    "Aperturas no autorizadas",
    "Acumulación de tardanzas",
    "Incumplimiento de procedimientos",
    "Otro",
]

# Gravedad sugerida por causal, para preseleccionar el tipo de llamado.
GRAVEDAD_CAUSAL = {
    "Errores de tasación":            "Escrito",
    "Aperturas no autorizadas":       "Suspensión",
    "Errores en cierre de caja":      "Escrito",
    "Errores de gestión":             "Verbal",
    "Mala atención al cliente":       "Verbal",
    "Acumulación de tardanzas":       "Verbal",
    "Incumplimiento de procedimientos": "Verbal",
    "Otro":                            "Verbal",
}

# Estados que aún cuentan contra el cupo mensual de permisos
ESTADOS_VIGENTES = {EST_PEND_JEFE.lower(), EST_PEND_RRHH.lower(), EST_APROBADO.lower(),
                    "pendiente", "pendiente_aprobacion"}


# ── Seguridad: hashing de contraseñas ─────────────────────────────────────────
# PBKDF2-HMAC-SHA256 con salt aleatorio por usuario. Reemplaza al SHA-256 plano,
# que era vulnerable a tablas rainbow porque el mismo password siempre producía
# el mismo hash. Los hashes antiguos se siguen aceptando y se migran solos en el
# primer inicio de sesión correcto.
_PBKDF2_ITERACIONES = 200_000


def hash_password(password: str, salt: bytes | None = None) -> str:
    """Genera un hash PBKDF2 con salt. Formato: pbkdf2$<iter>$<salt_b64>$<hash_b64>"""
    if salt is None:
        salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, _PBKDF2_ITERACIONES)
    return "pbkdf2${}${}${}".format(
        _PBKDF2_ITERACIONES,
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(dk).decode("ascii"),
    )


def verify_password(password: str, almacenado: str) -> bool:
    """Verifica la contraseña contra el hash almacenado (nuevo o heredado)."""
    almacenado = str(almacenado or "").strip()
    if not almacenado:
        return False
    if almacenado.startswith("pbkdf2$"):
        try:
            _, iteraciones, salt_b64, hash_b64 = almacenado.split("$")
            salt = base64.b64decode(salt_b64)
            esperado = base64.b64decode(hash_b64)
            dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iteraciones))
            return hmac.compare_digest(dk, esperado)
        except Exception:
            return False
    # Hash heredado: SHA-256 sin salt
    heredado = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return hmac.compare_digest(heredado, almacenado)


def es_hash_heredado(almacenado: str) -> bool:
    """True si el hash usa el formato antiguo y conviene migrarlo."""
    return bool(almacenado) and not str(almacenado).strip().startswith("pbkdf2$")


def password_debil(password: str) -> str | None:
    """Devuelve un mensaje si la contraseña no cumple la política mínima, o None."""
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres."
    if password.isdigit():
        return "La contraseña no puede ser solo números."
    if password.isalpha():
        return "La contraseña debe incluir al menos un número o símbolo."
    comunes = {"12345678", "password", "quski2026", "contrasena", "administrador", "abc12345"}
    if password.lower() in comunes:
        return "Esa contraseña es demasiado común. Elige otra."
    return None




# ── Lectura del formulario "Conozca a su Empleado" en Excel ───────────────────
# El formulario tiene posiciones fijas, así que se leen por coordenada. Los
# campos de texto salen limpios; los de "marque con una x" son menos confiables
# porque la marca a veces va dentro de la etiqueta y a veces no se pone, así
# que se extraen igual pero SIEMPRE se presentan para revisión antes de guardar.

_COORD_KYE = {
    "apellido1": "C8",  "apellido2": "E8",  "nombres": "G8",  "Cedula": "L8",
    "Nacionalidad": "C11", "Ciudad_Nacimiento": "D11",
    "nac_dia": "E11", "nac_mes": "F11", "nac_anio": "G11",
    "hijos_txt": "E13",
    "pais": "C18", "Provincia": "D18", "Canton": "E18", "Parroquia": "F18",
    "Direccion": "G18", "Telefono_Domicilio": "J18", "Celular": "K18",
    "Email_Personal": "L18",
    "cy_ap1": "C22", "cy_ap2": "E22", "cy_nom": "G22", "Conyuge_Cedula": "L22",
    "Conyuge_Empresa": "D33", "Conyuge_Actividad": "F33",
    "Conyuge_Cargo": "L33", "Conyuge_Anios_Cargo": "M33",
    "cy_publico": "C33", "cy_privado": "C35", "cy_independiente": "C25",
    "Fecha_Formulario": "D93",
}

_CIVIL_COORD = {"H11": "Soltero(a)", "I11": "Casado(a)", "J11": "Divorciado(a)",
                "K11": "Viudo(a)", "L11": "Separado(a)", "M11": "Unión libre"}
_SEXO_COORD  = {"C14": "Femenino", "D14": "Masculino"}


def _celda(ws, coord: str) -> str:
    try:
        v = ws[coord].value
    except Exception:
        return ""
    return "" if v is None else str(v).strip()


def _tiene_marca(texto: str) -> bool:
    """Detecta la 'x' con que se marcan las casillas del formulario."""
    t = (texto or "").strip().lower()
    if not t:
        return False
    # La marca suele venir pegada a la etiqueta: "Femenino   x"
    return t.endswith(" x") or t.endswith("\tx") or t == "x" or " x " in f" {t} "


def _opcion_marcada(ws, mapa: dict) -> str:
    """Devuelve la etiqueta cuya casilla está marcada, o cadena vacía."""
    for coord, etiqueta in mapa.items():
        if _tiene_marca(_celda(ws, coord)):
            return etiqueta
    return ""


def _solo_numero(texto: str) -> str:
    import re
    m = re.findall(r"\d+", texto or "")
    return m[-1] if m else ""


def leer_formulario_kye(archivo) -> dict:
    """Extrae los datos de un formulario KYE en Excel.

    Devuelve {"datos": {...campos de KYE_Empleado...},
              "nombre_detectado": str,
              "avisos": [str]}   — avisos son los campos que no se pudieron leer
                                   y hay que confirmar a mano.
    """
    import openpyxl
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb["C_EMPLEADO"] if "C_EMPLEADO" in wb.sheetnames else wb[wb.sheetnames[0]]

    c = {k: _celda(ws, v) for k, v in _COORD_KYE.items()}
    avisos = []

    # Fecha de nacimiento, armada desde tres celdas separadas
    dia, mes, anio = (_solo_numero(c["nac_dia"]), _solo_numero(c["nac_mes"]),
                      _solo_numero(c["nac_anio"]))
    if dia and mes and anio:
        f_nac = f"{int(anio):04d}-{int(mes):02d}-{int(dia):02d}"
    else:
        f_nac = ""
        avisos.append("Fecha de nacimiento")

    civil = _opcion_marcada(ws, _CIVIL_COORD)
    if not civil:
        avisos.append("Estado civil (no se detectó la marca)")
    sexo = _opcion_marcada(ws, _SEXO_COORD)
    if not sexo:
        avisos.append("Sexo (no se detectó la marca)")

    hijos = _solo_numero(c["hijos_txt"])
    if not hijos:
        avisos.append("Número de hijos")

    nombre_completo = " ".join(x for x in (c["nombres"], c["apellido1"],
                                           c["apellido2"]) if x).strip()
    conyuge = " ".join(x for x in (c["cy_nom"], c["cy_ap1"], c["cy_ap2"]) if x).strip()

    # Relación laboral del cónyuge: la marca está junto a la etiqueta
    if _tiene_marca(c["cy_publico"]) or c["Conyuge_Empresa"]:
        relacion = "Empleado público" if _tiene_marca(c["cy_publico"]) else "Empleado privado"
    elif _tiene_marca(c["cy_privado"]):
        relacion = "Empleado privado"
    elif _tiene_marca(c["cy_independiente"]):
        relacion = "Independiente"
    else:
        relacion = ""

    trabaja = "Si" if (c["Conyuge_Empresa"] or c["Conyuge_Cargo"]) else "No"

    # Las declaraciones PEP se marcan con x sobre etiquetas SI/NO. Si no se
    # detecta marca no se asume nada: es un dato regulatorio, se confirma.
    pep_si = _tiene_marca(_celda(ws, "L40"))
    pep_no = _tiene_marca(_celda(ws, "M40"))
    fam_si = _tiene_marca(_celda(ws, "K46"))
    fam_no = _tiene_marca(_celda(ws, "L46"))
    es_pep = "Si" if pep_si else ("No" if pep_no else "")
    fam_pep = "Si" if fam_si else ("No" if fam_no else "")
    if not es_pep:
        avisos.append("Declaración PEP propia")
    if not fam_pep:
        avisos.append("Declaración PEP de familiares")

    datos = {
        "Cedula": c["Cedula"], "Fecha_Nacimiento": f_nac,
        "Nacionalidad": c["Nacionalidad"], "Ciudad_Nacimiento": c["Ciudad_Nacimiento"],
        "Estado_Civil": civil, "Sexo": sexo, "Num_Hijos": hijos or "0",
        "Provincia": c["Provincia"], "Canton": c["Canton"],
        "Parroquia": c["Parroquia"], "Direccion": c["Direccion"],
        "Telefono_Domicilio": c["Telefono_Domicilio"], "Celular": c["Celular"],
        "Email_Personal": c["Email_Personal"],
        "Conyuge_Nombre": conyuge, "Conyuge_Cedula": c["Conyuge_Cedula"],
        "Conyuge_Trabaja": trabaja, "Conyuge_Relacion": relacion,
        "Conyuge_Empresa": c["Conyuge_Empresa"],
        "Conyuge_Actividad": c["Conyuge_Actividad"],
        "Conyuge_Cargo": c["Conyuge_Cargo"],
        "Conyuge_Anios_Cargo": c["Conyuge_Anios_Cargo"],
        "Es_PEP": es_pep, "Familiar_PEP": fam_pep,
        "Fecha_Formulario": c["Fecha_Formulario"],
    }
    return {"datos": datos, "nombre_detectado": nombre_completo, "avisos": avisos}


# ── SheetsManager ─────────────────────────────────────────────────────────────
class SheetsManager:
    """Wrapper sobre gspread para el spreadsheet de Asistencia RRHH."""

    def __init__(self, credentials_source):
        """
        credentials_source: ruta a credentials.json (str)
                           ó dict con el contenido del JSON de service account.
        """
        if isinstance(credentials_source, dict):
            creds = Credentials.from_service_account_info(credentials_source, scopes=SCOPES)
        else:
            creds = Credentials.from_service_account_file(credentials_source, scopes=SCOPES)

        self.credenciales = creds          # necesarias para el servicio de Drive
        self._drive = None
        self.client = gspread.authorize(creds)
        self.spreadsheet = con_reintentos(self.client.open_by_key, SPREADSHEET_ID)
        # Caché de lecturas y último error de lectura, para no confundir un
        # fallo de la API con "no hay datos".
        self._cache = {}
        self.ultimo_error = None
        self._hojas_cache = None      # lista de pestañas, se pide una sola vez
        self._precarga_seg = 0        # cuándo se leyeron todas las hojas juntas

    # ── Helpers de bajo nivel ────────────────────────────────────────────────

    def _hojas(self, refrescar: bool = False) -> dict:
        """Pestañas del archivo, pedidas UNA sola vez por sesión.

        gspread consulta la lista de pestañas al servidor en cada llamada a
        worksheet(); con catorce hojas eso era la mitad del consumo de cuota.
        """
        if refrescar or getattr(self, "_hojas_cache", None) is None:
            self._hojas_cache = {ws.title: ws
                                 for ws in con_reintentos(self.spreadsheet.worksheets)}
        return self._hojas_cache

    def _sheet(self, name: str):
        hojas = self._hojas()
        if name not in hojas:
            hojas = self._hojas(refrescar=True)   # quizá se acaba de crear
        if name not in hojas:
            raise gspread.exceptions.WorksheetNotFound(name)
        return hojas[name]

    def _df_desde_valores(self, sheet_name: str, valores: list) -> pd.DataFrame:
        """Arma la tabla desde las celdas en crudo.

        Resuelve los dos casos que rompen la lectura estándar: encabezados
        repetidos —al segundo se le añade un sufijo— y encabezados en blanco,
        cuya columna se descarta. Así una hoja con una columna de sobra se
        sigue leyendo entera en vez de no leerse nada.
        """
        esperadas = HEADERS.get(sheet_name, [])
        if not valores:
            return pd.DataFrame(columns=list(esperadas) + [FILA_HOJA])
        cabecera = [str(c).strip() for c in valores[0]]
        usados, columnas = {}, []
        for nombre in cabecera:
            if not nombre:
                columnas.append(None)
                continue
            # Nombres equivalentes salvo mayúsculas se llevan al nombre oficial
            for exp in esperadas:
                if nombre.lower() == exp.lower():
                    nombre = exp
                    break
            if nombre in usados:
                usados[nombre] += 1
                nombre = f"{nombre}_{usados[nombre]}"
            else:
                usados[nombre] = 0
            columnas.append(nombre)
        registros = []
        for n_fila, fila in enumerate(valores[1:], start=2):   # 2 = primera fila de datos
            if not any(str(c).strip() for c in fila):
                continue                       # fila totalmente en blanco
            reg = {}
            for i, nombre in enumerate(columnas):
                if nombre is not None:
                    reg[nombre] = fila[i] if i < len(fila) else ""
            if reg:
                # El número real de fila en la hoja. Sin esto habría que
                # deducirlo de la posición, y basta una fila en blanco en medio
                # para que todas las escrituras siguientes caigan desplazadas.
                reg[FILA_HOJA] = n_fila
                registros.append(reg)
        if not registros:
            return pd.DataFrame(columns=esperadas)
        df = pd.DataFrame(registros)
        for exp in esperadas:
            if exp not in df.columns:
                df[exp] = ""
        return df

    def precargar(self, forzar: bool = False) -> bool:
        """Lee TODAS las hojas en una sola llamada a la API.

        Es la diferencia entre gastar catorce lecturas del minuto y gastar
        una. Google cobra por llamada, no por cantidad de datos, así que traer
        todo junto sale mucho más barato que pedir hoja por hoja.
        """
        ahora_seg = time.time()
        if not forzar and (ahora_seg - getattr(self, "_precarga_seg", 0)) < _CACHE_TTL:
            return True
        try:
            hojas = self._hojas()
            nombres = [n for n in HEADERS if n in hojas]
            if not nombres:
                return False
            resp = con_reintentos(self.spreadsheet.values_batch_get,
                                  [f"'{n}'" for n in nombres])
            rangos = resp.get("valueRanges", [])
            for nombre, rango in zip(nombres, rangos):
                df = self._df_desde_valores(nombre, rango.get("values", []) or [])
                self._cache[nombre] = (ahora_seg, df)
            self._precarga_seg = ahora_seg
            self.ultimo_error = None
            return True
        except Exception as e:
            self.ultimo_error = f"{type(e).__name__}: {e}"
            return False

    def _invalidar_cache(self, sheet_name: str = None):
        """Descarta la lectura guardada tras escribir en una hoja."""
        if sheet_name:
            self._cache.pop(sheet_name, None)
        else:
            self._cache.clear()
        # La próxima lectura debe volver a traer los datos de Google, no
        # servir la foto anterior: si no, quien acaba de guardar no ve lo suyo.
        self._precarga_seg = 0

    def get_df(self, sheet_name: str, usar_cache: bool = True) -> pd.DataFrame:
        """Lee la hoja y devuelve un DataFrame con las columnas esperadas.

        No pide la hoja suelta: pide TODAS de una vez y se queda con la que
        hace falta. Una pantalla que consulta seis hojas cuesta entonces una
        llamada, no seis, que es lo que agotaba la cuota del minuto.
        """
        if usar_cache:
            guardado = self._cache.get(sheet_name)
            if guardado and (time.time() - guardado[0]) < _CACHE_TTL:
                return guardado[1].copy()

        if self.precargar(forzar=not usar_cache):
            guardado = self._cache.get(sheet_name)
            if guardado:
                return guardado[1].copy()

        # La lectura por lotes no funcionó. Último intento con la hoja suelta,
        # salvo que el problema sea la cuota: ahí insistir solo la empeora.
        if es_error_cuota(Exception(self.ultimo_error or "")):
            return pd.DataFrame(columns=HEADERS[sheet_name])
        try:
            valores = con_reintentos(self._sheet(sheet_name).get_all_values)
            df = self._df_desde_valores(sheet_name, valores)
            self._cache[sheet_name] = (time.time(), df)
            self.ultimo_error = None
            return df.copy()
        except Exception as e:
            self.ultimo_error = f"{type(e).__name__}: {e}"
            return pd.DataFrame(columns=HEADERS[sheet_name])

    def append(self, sheet_name: str, row: list):
        """Agrega una fila al final de la hoja."""
        con_reintentos(self._sheet(sheet_name).append_row, row,
                       value_input_option="USER_ENTERED")
        self._invalidar_cache(sheet_name)

    def update_cell(self, sheet_name: str, row: int, col: int, value):
        """Actualiza una celda (row/col base 1, fila 1 = encabezado)."""
        con_reintentos(self._sheet(sheet_name).update_cell, row, col, value)
        self._invalidar_cache(sheet_name)

    def update_row(self, sheet_name: str, row_idx: int, data: list):
        """Actualiza la fila completa (row_idx base 1, fila 1 = encabezado)."""
        ncols = len(data)
        rango = f"A{row_idx}:{rowcol_to_a1(row_idx, ncols)}"
        con_reintentos(self._sheet(sheet_name).update, range_name=rango,
                       values=[data], value_input_option="USER_ENTERED")
        self._invalidar_cache(sheet_name)

    def update_campos(self, sheet_name: str, row_idx: int, campos: dict):
        """Actualiza varias columnas por nombre en una sola llamada a la API.
        campos: {nombre_columna: valor}"""
        headers = HEADERS.get(sheet_name, [])
        peticiones = []
        for nombre, valor in campos.items():
            if nombre not in headers:
                continue
            col_i = headers.index(nombre) + 1
            peticiones.append({
                "range": rowcol_to_a1(row_idx, col_i),
                "values": [[str(valor)]],
            })
        if peticiones:
            con_reintentos(self._sheet(sheet_name).batch_update, peticiones,
                           value_input_option="USER_ENTERED")
            self._invalidar_cache(sheet_name)

    def _fila_de(self, sheet_name: str, id_col: str, id_val: str) -> int:
        """Devuelve el índice base-1 de la fila cuyo id_col == id_val."""
        df = self.get_df(sheet_name)
        if df.empty or id_col not in df.columns:
            raise ValueError(f"Registro {id_val} no encontrado en {sheet_name}")
        col = df[id_col].astype(str).str.strip()
        objetivo = str(id_val).strip()
        coincide = df[col == objetivo]
        if coincide.empty:
            raise ValueError(f"Registro {id_val} no encontrado en {sheet_name}")
        return fila_de_hoja(df, coincide.index[0])

    def ensure_columns(self, sheet_name: str):
        """Agrega al final las columnas de HEADERS que falten en la hoja.
        Permite migrar hojas creadas con versiones anteriores sin perder datos."""
        esperadas = HEADERS.get(sheet_name, [])
        if not esperadas:
            return
        try:
            sheet = self._sheet(sheet_name)
        except Exception:
            return
        try:
            # El encabezado sale de la precarga cuando ya se hizo, para no
            # gastar una llamada por hoja solo en mirar la primera fila.
            guardado = self._cache.get(sheet_name)
            if guardado is not None:
                actuales = [c for c in guardado[1].columns]
            else:
                actuales = [str(c).strip() for c in sheet.row_values(1)]
        except Exception:
            actuales = []
        if not actuales:
            sheet.update(range_name=f"A1:{rowcol_to_a1(1, len(esperadas))}",
                         values=[esperadas], value_input_option="USER_ENTERED")
            return
        faltantes = [c for c in esperadas if c not in actuales]
        if not faltantes:
            return
        nuevos = actuales + faltantes
        if sheet.col_count < len(nuevos):
            sheet.add_cols(len(nuevos) - sheet.col_count)
        sheet.update(range_name=f"A1:{rowcol_to_a1(1, len(nuevos))}",
                     values=[nuevos], value_input_option="USER_ENTERED")

    def migrar_esquema(self):
        """Asegura que las hojas tengan las columnas que el sistema espera.

        Es idempotente: si ya están, no escribe nada. Empieza por una lectura
        por lotes para saber qué encabezado tiene cada hoja con una sola
        llamada, en vez de una por hoja.
        """
        self.precargar()
        for hoja in ("Permisos", "Vacaciones", "Empleados", "Asistencia"):
            try:
                self.ensure_columns(hoja)
            except Exception:
                pass

    # ── Configuración ────────────────────────────────────────────────────────

    def get_config(self) -> dict:
        df = self.get_df("Configuracion")
        if df.empty or "Key" not in df.columns:
            return CONFIG_DEFAULTS.copy()
        cfg = {}
        for _, row in df.iterrows():
            k = str(row["Key"]).strip()
            v = str(row["Valor"]).strip() if row["Valor"] is not None else ""
            # Ignorar vacíos, "nan" o "None" para que manden los valores por defecto
            if k and k not in ("nan", "None") and v and v not in ("nan", "None"):
                cfg[k] = v
        return {**CONFIG_DEFAULTS, **cfg}

    def save_all_config(self, config: dict):
        """Guarda toda la configuración de una vez: limpia la hoja y escribe
        la fila de encabezado + todas las claves/valores.  Más robusto que
        llamar a set_config N veces porque garantiza que A1 siempre sea el
        encabezado correcto y no quedan filas huérfanas."""
        sheet = self._sheet("Configuracion")
        rows = [["Key", "Valor"]] + [[k, str(v)] for k, v in config.items()]
        sheet.clear()
        self._invalidar_cache("Configuracion")
        result = sheet.update(range_name=f"A1:B{len(rows)}", values=rows,
                              value_input_option="USER_ENTERED")
        updated = (result or {}).get("updatedRows", 0)
        if updated and updated < len(rows):
            raise RuntimeError(
                f"Escritura incompleta: se esperaban {len(rows)} filas, "
                f"Google Sheets reportó {updated}."
            )

    def set_config(self, key: str, valor: str):
        """Guarda o actualiza una sola clave de configuración.
        Garantiza que la hoja tenga encabezado Key/Valor en la fila 1."""
        sheet = self._sheet("Configuracion")
        all_values = sheet.get_all_values()
        if not all_values:
            sheet.update(range_name="A1:B1", values=[["Key", "Valor"]],
                         value_input_option="USER_ENTERED")
        elif [str(v).strip() for v in all_values[0]] != ["Key", "Valor"]:
            sheet.insert_row(["Key", "Valor"], 1, value_input_option="USER_ENTERED")

        df = self.get_df("Configuracion")
        if not df.empty and "Key" in df.columns:
            keys = df["Key"].astype(str).tolist()
        else:
            keys = []
        if key in keys:
            row_idx = fila_de_hoja(df, df.index[keys.index(key)])
            self.update_row("Configuracion", row_idx, [key, valor])
        else:
            self.append("Configuracion", [key, valor])

    # ── Empleados ────────────────────────────────────────────────────────────

    def get_empleados(self) -> pd.DataFrame:
        return self.get_df("Empleados")

    def next_id_empleado(self) -> str:
        df = self.get_empleados()
        if df.empty or "ID_Empleado" not in df.columns:
            return "EMP001"
        ids = df["ID_Empleado"].astype(str).str.strip()
        ids = ids[ids != ""]
        if ids.empty:
            return "EMP001"
        nums = ids.str.extract(r"(\d+)$")[0].dropna().astype(int)
        next_n = int(nums.max()) + 1 if not nums.empty else 1
        if ids.str.match(r"^\d+$").all():
            return str(next_n)
        prefix_s = ids.str.extract(r"^([A-Za-z_]+)(\d+)$")
        if prefix_s is not None and not prefix_s.dropna().empty:
            prefix = prefix_s[0].dropna().iloc[0]
            digits = len(prefix_s[1].dropna().iloc[0])
            return f"{prefix}{next_n:0{digits}d}"
        return f"EMP{next_n:03d}"

    @staticmethod
    def _fila_empleado(emp_id, nombre, email, area, email_jefe, hora_inicio, hora_fin,
                       fecha_ingreso, dias_tomados_inicial, modalidad, dia_oficina) -> list:
        """Arma la fila por NOMBRE de columna, no por posición.

        Antes se pasaba una lista en orden fijo; al agregar columnas al final,
        una edición escribía una fila más corta que la hoja y dejaba los campos
        nuevos con el valor viejo sin que nadie se diera cuenta.
        """
        cols = HEADERS["Empleados"]
        fila = [""] * len(cols)
        for c, v in (("ID_Empleado", emp_id), ("Nombre", nombre), ("Email", email),
                     ("Area", area), ("Email_Jefe", email_jefe),
                     ("Horario_Inicio", hora_inicio), ("Horario_Fin", hora_fin),
                     ("Fecha_Ingreso", fecha_ingreso),
                     ("Dias_Tomados_Inicial", dias_tomados_inicial),
                     ("Modalidad", modalidad), ("Dia_Oficina", dia_oficina)):
            fila[cols.index(c)] = v
        return fila

    def agregar_empleado(self, nombre, email, area, email_jefe, hora_inicio, hora_fin,
                         fecha_ingreso="", dias_tomados_inicial=0,
                         modalidad=MOD_PRESENCIAL, dia_oficina="") -> str:
        emp_id = self.next_id_empleado()
        self.append("Empleados", self._fila_empleado(
            emp_id, nombre, email, area, email_jefe, hora_inicio, hora_fin,
            fecha_ingreso, dias_tomados_inicial, modalidad, dia_oficina))
        return emp_id

    def actualizar_empleado(self, emp_id: str, nombre, email, area, email_jefe,
                            hora_inicio, hora_fin, fecha_ingreso="", dias_tomados_inicial=0,
                            modalidad=MOD_PRESENCIAL, dia_oficina=""):
        row_idx = self._fila_de("Empleados", "ID_Empleado", emp_id)
        self.update_row("Empleados", row_idx, self._fila_empleado(
            emp_id, nombre, email, area, email_jefe, hora_inicio, hora_fin,
            fecha_ingreso, dias_tomados_inicial, modalidad, dia_oficina))

    def set_datos_vacaciones(self, emp_id: str, fecha_ingreso, dias_tomados_inicial):
        """Carga inicial: fecha de ingreso y días ya tomados antes del sistema."""
        row_idx = self._fila_de("Empleados", "ID_Empleado", emp_id)
        self.update_campos("Empleados", row_idx, {
            "Fecha_Ingreso": fecha_ingreso,
            "Dias_Tomados_Inicial": dias_tomados_inicial,
        })

    # ── Jerarquía: jefe inmediato y subordinados ─────────────────────────────

    def get_empleado(self, id_empleado: str) -> dict | None:
        """Devuelve la fila del empleado como dict, o None."""
        df = self.get_empleados()
        if df.empty or "ID_Empleado" not in df.columns:
            return None
        mask = df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()
        rows = df[mask]
        return rows.iloc[0].to_dict() if not rows.empty else None

    def get_empleado_por_email(self, email: str) -> dict | None:
        """Busca un empleado por su correo (case-insensitive)."""
        df = self.get_empleados()
        if df.empty or "Email" not in df.columns or not str(email).strip():
            return None
        mask = df["Email"].astype(str).str.strip().str.lower() == str(email).strip().lower()
        rows = df[mask]
        return rows.iloc[0].to_dict() if not rows.empty else None

    def get_email_empleado(self, id_empleado: str) -> str:
        """Devuelve el email del empleado o cadena vacía si no se encuentra."""
        emp = self.get_empleado(id_empleado)
        return str(emp.get("Email", "")).strip() if emp else ""

    def get_email_jefe(self, id_empleado: str) -> str:
        """Devuelve el correo del jefe inmediato del empleado, o cadena vacía."""
        emp = self.get_empleado(id_empleado)
        if not emp:
            return ""
        jefe = str(emp.get("Email_Jefe", "")).strip()
        return "" if jefe.lower() in ("", "nan", "none") else jefe

    def get_nombre_empleado(self, id_empleado: str) -> str:
        emp = self.get_empleado(id_empleado)
        return str(emp.get("Nombre", id_empleado)) if emp else str(id_empleado)

    def get_subordinados(self, email_jefe: str) -> pd.DataFrame:
        """Empleados cuyo Email_Jefe coincide con el correo dado."""
        df = self.get_empleados()
        if df.empty or "Email_Jefe" not in df.columns or not str(email_jefe).strip():
            return pd.DataFrame(columns=HEADERS["Empleados"])
        mask = df["Email_Jefe"].astype(str).str.strip().str.lower() == str(email_jefe).strip().lower()
        return df[mask]

    def es_jefe(self, email: str) -> bool:
        """True si alguien reporta a este correo."""
        return not self.get_subordinados(email).empty

    def ids_subordinados(self, email_jefe: str) -> list:
        df = self.get_subordinados(email_jefe)
        if df.empty:
            return []
        return df["ID_Empleado"].astype(str).str.strip().tolist()

    # ── Asistencia ───────────────────────────────────────────────────────────

    def get_asistencia(self, fecha: str = None) -> pd.DataFrame:
        df = self.get_df("Asistencia")
        if fecha and not df.empty and "Fecha" in df.columns:
            df = df[df["Fecha"].astype(str) == fecha]
        return df

    def ya_registro_entrada(self, id_empleado: str, fecha: str) -> bool:
        df = self.get_asistencia(fecha)
        if df.empty or "ID_Empleado" not in df.columns:
            return False
        return (str(id_empleado).strip()
                in df["ID_Empleado"].astype(str).str.strip().values)

    def estado_jornada(self, id_empleado: str, fecha: str) -> dict:
        """Qué tiene marcado hoy esta persona: entrada, salida, ambas o nada.

        Permite que la pantalla de asistencia abra donde la persona la
        necesita, en vez de dejarla siempre frente al botón de entrada.
        """
        vacio = {"entrada": "", "salida": "", "estado": "", "completa": False}
        df = self.get_asistencia(fecha)
        if df.empty or "ID_Empleado" not in df.columns:
            return vacio
        fila = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
        if fila.empty:
            return vacio
        r = fila.iloc[-1]
        ent = str(r.get("Hora_Entrada", "") or "").strip()
        sal = str(r.get("Hora_Salida", "") or "").strip()
        return {"entrada": ent, "salida": sal,
                "estado": str(r.get("Estado", "") or "").strip(),
                "completa": bool(ent and sal)}

    def minutos_de_atraso(self, hora, config: dict) -> int:
        """Minutos de atraso que generaría marcar la entrada a esa hora.

        Se calcula igual que registrar_entrada, para poder advertirle a la
        persona ANTES de que pulse el botón cuánto atraso va a quedar en su
        expediente.
        """
        try:
            inicio = str(config.get("Horario_Inicio", "09:00"))[:5]
            tol = int(float(config.get("Tolerancia_Minutos", 0) or 0))
            fmt = "%H:%M"
            d = (datetime.strptime(str(hora)[:5], fmt)
                 - datetime.strptime(inicio, fmt)).total_seconds() / 60
            return int(max(0, d - tol))
        except Exception:
            return 0

    def registrar_entrada(self, id_empleado: str, nombre: str, hora_entrada: str,
                          config: dict, dispositivo: str = "") -> tuple:
        """Registra la hora de entrada, una sola vez al día.

        Devuelve (estado, minutos_atraso). Si la persona YA tiene entrada hoy,
        no escribe nada y devuelve la que ya existía: la defensa contra el
        registro duplicado vive aquí, no en la pantalla, porque la pantalla
        consulta una caché que puede tener hasta minuto y medio de atraso y
        cualquiera que abra el aplicativo dos veces la esquiva.
        """
        fecha = hoy_local(config).strftime("%Y-%m-%d")

        # Relectura SIN caché justo antes de escribir. Es una llamada más a
        # Google, y vale la pena: evita el registro doble, que ensucia el
        # expediente disciplinario de la persona.
        previa = self.entrada_de(id_empleado, fecha, usar_cache=False)
        if previa:
            return previa["estado"] or "A_Tiempo", int(previa["minutos_atraso"] or 0)

        minutos_atraso = self.minutos_de_atraso(hora_entrada, config)
        estado = "Tardanza" if minutos_atraso > 0 else "A_Tiempo"

        cols = HEADERS["Asistencia"]
        fila = [""] * len(cols)
        for nombre_col, valor in (("Fecha", fecha), ("ID_Empleado", id_empleado),
                                  ("Nombre", nombre), ("Hora_Entrada", hora_entrada),
                                  ("Estado", estado),
                                  ("Minutos_Atraso", int(minutos_atraso)),
                                  ("Dispositivo_Entrada", dispositivo)):
            fila[cols.index(nombre_col)] = valor
        self.append("Asistencia", fila)
        return estado, int(minutos_atraso)

    def entrada_de(self, id_empleado: str, fecha: str, usar_cache: bool = True) -> dict | None:
        """El registro de asistencia de esa persona ese día, si existe."""
        df = self.get_df("Asistencia", usar_cache=usar_cache)
        if df.empty or "ID_Empleado" not in df.columns:
            return None
        m = df[(df["Fecha"].astype(str).str.strip() == str(fecha).strip())
               & (df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip())]
        if m.empty:
            return None
        r = m.iloc[0]
        return {"fila": fila_de_hoja(df, m.index[0]),
                "entrada": str(r.get("Hora_Entrada", "") or "").strip(),
                "salida": str(r.get("Hora_Salida", "") or "").strip(),
                "estado": str(r.get("Estado", "") or "").strip(),
                "minutos_atraso": self._a_numero(r.get("Minutos_Atraso"), 0),
                "duplicadas": len(m)}

    def entradas_duplicadas(self, dias: int = 60) -> list:
        """Días en que una misma persona quedó con más de un registro.

        Cada duplicado infla el conteo de atrasos del mes y deja una salida sin
        pareja, así que conviene limpiarlos.
        """
        df = self.get_df("Asistencia")
        if df.empty or "ID_Empleado" not in df.columns:
            return []
        corte = hoy_local() - timedelta(days=dias)
        salida = []
        for (fecha, emp), g in df.groupby([df["Fecha"].astype(str).str.strip(),
                                           df["ID_Empleado"].astype(str).str.strip()]):
            if len(g) < 2:
                continue
            f = self._a_fecha(fecha)
            if not f or f < corte:
                continue
            filas = [{"fila": fila_de_hoja(df, i),
                      "entrada": str(g.at[i, "Hora_Entrada"] or "").strip(),
                      "salida": str(g.at[i, "Hora_Salida"] or "").strip(),
                      "estado": str(g.at[i, "Estado"] or "").strip(),
                      "atraso": self._a_numero(g.at[i, "Minutos_Atraso"], 0),
                      "dispositivo": str(g.at[i, "Dispositivo_Entrada"] or "").strip()
                      if "Dispositivo_Entrada" in g.columns else ""}
                     for i in g.index]
            filas.sort(key=lambda x: x["entrada"] or "99:99")
            salida.append({"fecha": fecha, "id_empleado": emp,
                           "nombre": str(g.iloc[0].get("Nombre", "")).strip(),
                           "n": len(g), "filas": filas,
                           # Se conserva la PRIMERA marca del día, que es la
                           # real; las posteriores son reaperturas del
                           # aplicativo y son las que hay que quitar.
                           "conservar": filas[0]["fila"],
                           "sobrantes": [f["fila"] for f in filas[1:]]})
        salida.sort(key=lambda x: (x["fecha"], x["nombre"]), reverse=True)
        return salida

    def limpiar_entrada_duplicada(self, fila: int):
        """Vacía una fila duplicada en vez de borrarla.

        Borrar la fila correría todas las de abajo y desalinearía cualquier
        referencia que alguien tenga anotada; dejarla en blanco la saca de
        todos los conteos sin mover nada.
        """
        cols = HEADERS["Asistencia"]
        con_reintentos(self._sheet("Asistencia").batch_clear,
                       [f"A{fila}:{rowcol_to_a1(fila, len(cols))}"])
        self._invalidar_cache("Asistencia")

    def registrar_salida(self, id_empleado: str, fecha: str, hora_salida: str,
                         observaciones: str = "", config: dict | None = None,
                         dispositivo: str = "") -> dict:
        """Registra la hora de salida.

        La salida NO genera atrasos ni horas extra: salir después del horario
        solo deja constancia de la hora. Lo único que se evalúa es si la salida
        fue ANTICIPADA, para poder avisar al jefe inmediato.

        Devuelve {"anticipada", "minutos_antes", "horario_fin", "hora_salida"}.
        """
        df = self.get_df("Asistencia")
        if df.empty:
            raise ValueError("No hay registros de asistencia")
        mask = (df["Fecha"].astype(str) == fecha) & (df["ID_Empleado"].astype(str) == str(id_empleado))
        rows = df[mask]
        if rows.empty:
            raise ValueError(f"No se encontró entrada registrada para {id_empleado} el {fecha}")
        row_idx = fila_de_hoja(df, rows.index[0])
        campos = {"Hora_Salida": hora_salida}
        if observaciones:
            campos["Observaciones"] = observaciones
        if dispositivo:
            campos["Dispositivo_Salida"] = dispositivo
        self.update_campos("Asistencia", row_idx, campos)

        permiso = self.tiene_permiso_vigente(id_empleado, fecha)
        return self.evaluar_salida(hora_salida, config or {}, permiso)

    def tiene_permiso_vigente(self, id_empleado: str, fecha: str) -> dict | None:
        """Devuelve el permiso aprobado o en trámite del empleado en esa fecha.

        Sirve para no marcar como salida anticipada a quien tenía permiso: se
        fue antes porque estaba autorizado.
        """
        df = self.get_permisos()
        if df.empty or "ID_Empleado" not in df.columns:
            return None
        f = self._a_fecha(fecha)
        for _, r in df.iterrows():
            if str(r.get("ID_Empleado", "")).strip() != str(id_empleado).strip():
                continue
            if self._a_fecha(r.get("Fecha")) != f:
                continue
            estado = str(r.get("Estado", "")).strip()
            if estado in (EST_APROBADO, EST_PEND_JEFE, EST_PEND_RRHH,
                          "Pendiente", "Pendiente_Aprobacion"):
                return {"id": str(r.get("ID_Permiso", "")).strip(),
                        "horas": self._a_numero(r.get("Horas_Solicitadas"), 0),
                        "motivo": str(r.get("Motivo", "")).strip(),
                        "estado": estado,
                        "aprobado": estado == EST_APROBADO}
        return None

    def evaluar_salida(self, hora_salida: str, config: dict,
                       permiso: dict | None = None) -> dict:
        """Compara la hora de salida contra el horario de fin.

        Solo interesa la salida anticipada. Salir más tarde del horario no
        produce atraso ni hora extra: se registra la hora y nada más.
        """
        horario_fin = str(config.get("Horario_Fin", "17:30"))[:5]
        tolerancia = int(self._a_numero(config.get("Tolerancia_Salida_Minutos"), 0))
        try:
            fin = datetime.strptime(horario_fin, "%H:%M")
            sal = datetime.strptime(str(hora_salida)[:5], "%H:%M")
        except ValueError:
            return {"anticipada": False, "minutos_antes": 0,
                    "horario_fin": horario_fin, "hora_salida": hora_salida}
        minutos_antes = int((fin - sal).total_seconds() / 60)
        # Con permiso vigente ese día la salida temprana está autorizada: no es
        # una salida anticipada y no se avisa a nadie.
        anticipada = minutos_antes > tolerancia and permiso is None
        return {"anticipada": anticipada,
                "minutos_antes": max(0, minutos_antes),
                "horario_fin": horario_fin, "hora_salida": str(hora_salida)[:5],
                "permiso": permiso}

    def salidas_pendientes(self, dias_atras: int = 7, config: dict | None = None) -> list:
        """Jornadas de días anteriores que quedaron sin marcar la salida.

        Solo mira días ya cerrados (no el de hoy, que sigue en curso) y omite
        las que ya tienen aviso enviado, para no repetir correos.
        """
        df = self.get_df("Asistencia")
        if df.empty:
            return []
        hoy = hoy_local(config)
        desde = hoy - timedelta(days=dias_atras)
        pendientes = []
        for pos, (_, r) in enumerate(df.iterrows()):
            f = self._a_fecha(r.get("Fecha"))
            if not f or not (desde <= f < hoy):
                continue
            if str(r.get("Estado", "")).strip().lower() == "ausente":
                continue
            if not str(r.get("Hora_Entrada", "")).strip():
                continue
            if str(r.get("Hora_Salida", "")).strip():
                continue
            if str(r.get("Aviso_Salida", "")).strip():
                continue        # ya se avisó
            pendientes.append({
                # Fila real de la hoja, no la posición en el DataFrame.
                "fila": int(r.get(FILA_HOJA) or (pos + 2)),
                "fecha": f,
                "id_empleado": str(r.get("ID_Empleado", "")).strip(),
                "nombre": str(r.get("Nombre", "")).strip(),
                "hora_entrada": str(r.get("Hora_Entrada", "")).strip(),
            })
        return pendientes

    def marcar_aviso_salida(self, fila: int, texto: str):
        """Deja constancia de que ya se avisó por esta salida sin marcar."""
        self.update_campos("Asistencia", fila, {"Aviso_Salida": texto})

    def marcar_ausencia(self, id_empleado: str, nombre: str, fecha: str, motivo: str = ""):
        cols = HEADERS["Asistencia"]
        fila = [""] * len(cols)
        for c, v in (("Fecha", fecha), ("ID_Empleado", id_empleado), ("Nombre", nombre),
                     ("Estado", "Ausente"), ("Minutos_Atraso", 0),
                     ("Observaciones", motivo)):
            fila[cols.index(c)] = v
        self.append("Asistencia", fila)

    # ── Permisos ─────────────────────────────────────────────────────────────

    def get_permisos(self) -> pd.DataFrame:
        return self.get_df("Permisos")

    def horas_permiso_usadas_mes(self, id_empleado: str, año_mes: str) -> float:
        """Horas de permiso ya comprometidas en el mes (aprobadas o en trámite)."""
        df = self.get_permisos()
        if df.empty or "ID_Empleado" not in df.columns:
            return 0.0
        mask = (df["ID_Empleado"].astype(str) == str(id_empleado)) & \
               (df["Fecha"].astype(str).str.startswith(año_mes)) & \
               (df["Estado"].astype(str).str.lower().isin(ESTADOS_VIGENTES))
        total = pd.to_numeric(df[mask]["Horas_Solicitadas"], errors="coerce").fillna(0).sum()
        return float(total)

    def _next_id(self, sheet_name: str, id_col: str, prefijo: str, ancho: int = 4) -> str:
        df = self.get_df(sheet_name)
        if df.empty or id_col not in df.columns:
            return f"{prefijo}{1:0{ancho}d}"
        nums = pd.to_numeric(df[id_col].astype(str).str.extract(r"(\d+)$")[0], errors="coerce").dropna()
        next_n = int(nums.max()) + 1 if not nums.empty else 1
        return f"{prefijo}{next_n:0{ancho}d}"

    def solicitar_permiso(self, id_empleado: str, fecha: str, horas: float,
                          motivo: str, config: dict) -> dict:
        """Crea la solicitud de permiso. Siempre requiere aprobación del jefe
        inmediato (si está asignado) y luego de RRHH.

        Devuelve {"id", "estado", "excede_cupo", "email_jefe", "horas_usadas"}.
        """
        año_mes = fecha[:7]
        usadas = self.horas_permiso_usadas_mes(id_empleado, año_mes)
        limite = float(config.get("Horas_Permiso_Mensual", 3))
        excede = (usadas + horas) > limite

        email_jefe = self.get_email_jefe(id_empleado)
        estado = EST_PEND_JEFE if email_jefe else EST_PEND_RRHH

        perm_id = self._next_id("Permisos", "ID_Permiso", "PERM")
        fila = [perm_id, fecha, id_empleado, horas, motivo, estado,
                usadas + horas, "", "", "", "", "", "", ""]
        self.append("Permisos", fila)
        return {
            "id": perm_id,
            "estado": estado,
            "excede_cupo": excede,
            "email_jefe": email_jefe,
            "horas_usadas": usadas + horas,
            "limite": limite,
        }

    def aprobar_permiso_jefe(self, perm_id: str, aprobador: str):
        """Aprobación del jefe inmediato. Pasa la solicitud a RRHH."""
        row = self._fila_de("Permisos", "ID_Permiso", perm_id)
        self.update_campos("Permisos", row, {
            "Estado":           EST_PEND_RRHH,
            "Aprobado_Jefe":    aprobador,
            "Fecha_Aprob_Jefe": ahora_local().strftime("%Y-%m-%d %H:%M"),
        })

    def aprobar_permiso_rrhh(self, perm_id: str, aprobador: str):
        """Aprobación final de RRHH."""
        row = self._fila_de("Permisos", "ID_Permiso", perm_id)
        self.update_campos("Permisos", row, {
            "Estado":           EST_APROBADO,
            "Aprobado_RRHH":    aprobador,
            "Fecha_Aprob_RRHH": ahora_local().strftime("%Y-%m-%d %H:%M"),
            "Aprobado_Por":     aprobador,
        })

    def rechazar_permiso(self, perm_id: str, rechazado_por: str, motivo: str):
        row = self._fila_de("Permisos", "ID_Permiso", perm_id)
        self.update_campos("Permisos", row, {
            "Estado":         EST_RECHAZADO,
            "Motivo_Rechazo": motivo,
            "Rechazado_Por":  rechazado_por,
        })

    # Compatibilidad con la versión anterior
    def aprobar_permiso(self, perm_id: str, aprobado_por: str):
        self.aprobar_permiso_rrhh(perm_id, aprobado_por)

    def permisos_pendientes_jefe(self, email_jefe: str) -> pd.DataFrame:
        """Permisos esperando la firma de este jefe."""
        ids = self.ids_subordinados(email_jefe)
        df = self.get_permisos()
        if df.empty or not ids:
            return pd.DataFrame(columns=HEADERS["Permisos"])
        mask = df["ID_Empleado"].astype(str).str.strip().isin(ids) & \
               (df["Estado"].astype(str) == EST_PEND_JEFE)
        return df[mask]

    def permisos_pendientes_rrhh(self) -> pd.DataFrame:
        df = self.get_permisos()
        if df.empty:
            return pd.DataFrame(columns=HEADERS["Permisos"])
        # Incluye el estado antiguo para no perder solicitudes previas
        mask = df["Estado"].astype(str).isin([EST_PEND_RRHH, "Pendiente_Aprobacion"])
        return df[mask]

    # ── Vacaciones ───────────────────────────────────────────────────────────

    def get_vacaciones(self) -> pd.DataFrame:
        return self.get_df("Vacaciones")

    # ── Panel semanal ────────────────────────────────────────────────────────

    def feriados(self, config: dict | None = None) -> set:
        """Fechas marcadas como feriado en Configuración."""
        crudo = str((config or {}).get("Feriados", "") or "")
        fechas = set()
        for trozo in crudo.replace(";", ",").replace("\n", ",").split(","):
            f = self._a_fecha(trozo.strip())
            if f:
                fechas.add(f)
        return fechas

    def dias_laborables(self, d0, d1, config: dict | None = None) -> list:
        """Días de lunes a viernes del período, sin los feriados."""
        fer = self.feriados(config)
        dias, actual = [], d0
        while actual <= d1:
            if actual.weekday() < 5 and actual not in fer:
                dias.append(actual)
            actual += timedelta(days=1)
        return dias

    @staticmethod
    def modalidad_de(emp) -> str:
        """Modalidad de la persona; Presencial cuando la ficha no la declara."""
        try:
            v = str(emp.get("Modalidad", "") or "").strip().title()
        except Exception:
            v = ""
        return v if v in MODALIDADES else MOD_PRESENCIAL

    @staticmethod
    def _horas_entre(entrada: str, salida: str) -> float:
        """Horas entre dos marcas HH:MM. Devuelve 0 si alguna falta o no cuadra."""
        try:
            e = datetime.strptime(str(entrada)[:5], "%H:%M")
            sal = datetime.strptime(str(salida)[:5], "%H:%M")
        except (ValueError, TypeError):
            return 0.0
        h = (sal - e).total_seconds() / 3600
        return round(h, 2) if 0 < h < 24 else 0.0

    def resumen_periodo(self, fecha_ini, fecha_fin, config: dict | None = None) -> list:
        """Cumplimiento de cada persona en el período, listo para el panel.

        Lee las cuatro hojas una sola vez y cruza en memoria: asistencia contra
        los días que a esa persona le tocaba trabajar, descontando vacaciones
        aprobadas para que un descanso autorizado no aparezca como falta.

        Devuelve una lista de diccionarios, uno por empleado.
        """
        d0, d1 = self._a_fecha(fecha_ini), self._a_fecha(fecha_fin)
        if not d0 or not d1 or d1 < d0:
            return []
        config = config or {}
        laborables = self.dias_laborables(d0, d1, config)

        emps = self.get_empleados()
        if emps.empty:
            return []
        asis = self.get_asistencia()
        vac  = self.get_vacaciones()
        perm = self.get_permisos()

        # Vacaciones aprobadas: qué días quedan cubiertos para cada persona
        cubiertos = {}
        if not vac.empty and "ID_Empleado" in vac.columns:
            for _, v in vac.iterrows():
                if str(v.get("Estado", "")).strip() != EST_APROBADO:
                    continue
                vi, vf = self._a_fecha(v.get("Fecha_Inicio")), self._a_fecha(v.get("Fecha_Fin"))
                if not vi or not vf:
                    continue
                emp = str(v.get("ID_Empleado", "")).strip()
                dia = vi
                while dia <= vf:
                    if d0 <= dia <= d1:
                        cubiertos.setdefault(emp, set()).add(dia)
                    dia += timedelta(days=1)

        # Permisos aprobados: horas del período (no quitan el día, lo acortan)
        horas_permiso = {}
        if not perm.empty and "ID_Empleado" in perm.columns:
            for _, p in perm.iterrows():
                if str(p.get("Estado", "")).strip() != EST_APROBADO:
                    continue
                f = self._a_fecha(p.get("Fecha"))
                if not f or not (d0 <= f <= d1):
                    continue
                emp = str(p.get("ID_Empleado", "")).strip()
                horas_permiso[emp] = horas_permiso.get(emp, 0.0) + \
                    self._a_numero(p.get("Horas_Solicitadas"), 0)

        # Marcas de asistencia del período, indexadas por persona y día
        marcas = {}
        if not asis.empty and "ID_Empleado" in asis.columns:
            for _, r in asis.iterrows():
                f = self._a_fecha(r.get("Fecha"))
                if not f or not (d0 <= f <= d1):
                    continue
                emp = str(r.get("ID_Empleado", "")).strip()
                marcas.setdefault(emp, {})[f] = {
                    "entrada": str(r.get("Hora_Entrada", "") or "").strip(),
                    "salida":  str(r.get("Hora_Salida", "") or "").strip(),
                    "estado":  str(r.get("Estado", "") or "").strip(),
                    "atraso":  self._a_numero(r.get("Minutos_Atraso"), 0),
                    "dispositivo": str(r.get("Dispositivo_Entrada", "") or "").strip()
                                   if "Dispositivo_Entrada" in asis.columns else "",
                }

        salida = []
        for _, emp in emps.iterrows():
            eid  = str(emp.get("ID_Empleado", "")).strip()
            if not eid:
                continue
            mod  = self.modalidad_de(emp)
            dias_vac = cubiertos.get(eid, set())
            esperados = [d for d in laborables if d not in dias_vac]

            # Antes de su fecha de ingreso no se le puede exigir nada
            ingreso = self._a_fecha(emp.get("Fecha_Ingreso"))
            if ingreso:
                esperados = [d for d in esperados if d >= ingreso]

            mias = marcas.get(eid, {})
            dias_con_marca = [d for d in esperados if mias.get(d, {}).get("entrada")]
            sin_marca      = [d for d in esperados if not mias.get(d, {}).get("entrada")]
            sin_salida     = [d for d, m in mias.items() if m["entrada"] and not m["salida"]]
            atrasos        = [d for d, m in mias.items() if m["estado"] == "Tardanza"]
            ausencias      = [d for d, m in mias.items() if m["estado"] == "Ausente"]
            horas = sum(self._horas_entre(m["entrada"], m["salida"]) for m in mias.values())

            # Día de oficina: solo aplica a quien tiene uno pactado. El sistema
            # sabe si HUBO MARCA ese día, no si la persona estuvo en la oficina.
            dia_of = str(emp.get("Dia_Oficina", "") or "").strip()
            oficina = None
            if mod == MOD_MIXTO and dia_of in DIAS_SEMANA:
                idx = DIAS_SEMANA.index(dia_of)
                toca = [d for d in esperados if d.weekday() == idx]
                if toca:
                    oficina = all(mias.get(d, {}).get("entrada") for d in toca)

            salida.append({
                "id_empleado": eid,
                "nombre": str(emp.get("Nombre", "") or "").strip(),
                "area": str(emp.get("Area", "") or "").strip(),
                "modalidad": mod,
                "dia_oficina": dia_of,
                "dias_esperados": len(esperados),
                "dias_marcados": len(dias_con_marca),
                "dias_sin_marca": sorted(sin_marca),
                "dias_vacaciones": len(dias_vac),
                "horas_permiso": round(horas_permiso.get(eid, 0.0), 1),
                "atrasos": len(atrasos),
                "minutos_atraso": int(sum(mias[d]["atraso"] for d in atrasos)),
                "ausencias": len(ausencias),
                "sin_salida": sorted(sin_salida),
                "horas_declaradas": round(horas, 1),
                "dispositivos": sorted({m["dispositivo"] for m in mias.values()
                                        if m["dispositivo"]}),
                "oficina_cumplida": oficina,
                "detalle": {d: mias[d] for d in sorted(mias)},
            })
        salida.sort(key=lambda x: x["nombre"])
        return salida

    def dias_habiles(self, fecha_ini: str, fecha_fin: str) -> int:
        d0 = datetime.strptime(fecha_ini, "%Y-%m-%d")
        d1 = datetime.strptime(fecha_fin, "%Y-%m-%d")
        dias = 0
        current = d0
        while current <= d1:
            if current.weekday() < 5:  # Lunes-Viernes
                dias += 1
            current += timedelta(days=1)
        return dias


    # ── Saldo de vacaciones ──────────────────────────────────────────────────
    # Regla aplicada: 15 días CALENDARIO por cada año de servicio; a partir del
    # quinto año se suma un día por cada año adicional, con techo de 30.
    # Los tres números salen de Configuración, así que RRHH puede ajustarlos
    # sin tocar el código.

    @staticmethod
    def _a_fecha(valor):
        """Convierte a date lo que venga de la hoja, tolerando varios formatos."""
        if valor is None:
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        txt = str(valor).strip()
        if not txt or txt.lower() in ("nan", "none", ""):
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(txt[:10], fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _a_numero(valor, por_defecto=0.0) -> float:
        txt = str(valor if valor is not None else "").strip().replace(",", ".")
        if not txt or txt.lower() in ("nan", "none"):
            return por_defecto
        try:
            return float(txt)
        except ValueError:
            return por_defecto

    def dias_calendario(self, fecha_ini: str, fecha_fin: str) -> int:
        """Días corridos incluyendo fines de semana y feriados: así se cuentan
        las vacaciones por ley, a diferencia de los días hábiles."""
        d0 = self._a_fecha(fecha_ini)
        d1 = self._a_fecha(fecha_fin)
        if not d0 or not d1:
            return 0
        return max(0, (d1 - d0).days + 1)

    def anios_servicio(self, fecha_ingreso, referencia=None) -> float:
        ini = self._a_fecha(fecha_ingreso)
        if not ini:
            return 0.0
        ref = referencia or hoy_local()
        return max(0.0, (ref - ini).days / 365.25)

    def dias_derecho_del_anio(self, n_anio: int, config: dict) -> int:
        """Días que corresponden por el n-ésimo año de servicio."""
        base  = int(self._a_numero(config.get("Dias_Vacaciones_Base"), 15))
        desde = int(self._a_numero(config.get("Anio_Inicio_Dia_Adicional"), 5))
        techo = int(self._a_numero(config.get("Max_Dias_Vacaciones"), 30))
        if n_anio < desde:
            return base
        return min(techo, base + (n_anio - desde) + 1)

    def dias_acumulados(self, fecha_ingreso, config: dict, referencia=None) -> float:
        """Días ganados desde el ingreso hasta hoy, incluyendo la parte
        proporcional del año en curso."""
        ini = self._a_fecha(fecha_ingreso)
        if not ini:
            return 0.0
        ref = referencia or hoy_local()
        if ref <= ini:
            return 0.0
        total, aniv, n = 0.0, ini, 1
        while True:
            try:
                sig = aniv.replace(year=aniv.year + 1)
            except ValueError:          # 29 de febrero
                sig = aniv.replace(year=aniv.year + 1, day=28)
            dias_n = self.dias_derecho_del_anio(n, config)
            if sig <= ref:
                total += dias_n
                aniv, n = sig, n + 1
            else:
                largo = (sig - aniv).days
                if largo > 0:
                    total += dias_n * ((ref - aniv).days / largo)
                break
        return round(total, 1)

    def _dias_de_solicitud(self, fila) -> float:
        """Días calendario de una solicitud. Las solicitudes anteriores a este
        cambio solo guardaron días hábiles, así que se recalculan por fechas."""
        v = self._a_numero(fila.get("Dias_Calendario"), 0.0)
        if v > 0:
            return v
        return float(self.dias_calendario(fila.get("Fecha_Inicio"),
                                          fila.get("Fecha_Fin")))

    def dias_vacaciones_usados(self, id_empleado: str) -> dict:
        """Días ya usados: la carga inicial más lo aprobado y lo en trámite."""
        emp = self.get_empleado(id_empleado) or {}
        inicial = self._a_numero(emp.get("Dias_Tomados_Inicial"), 0.0)
        aprobados = en_tramite = 0.0
        df = self.get_vacaciones()
        if not df.empty and "ID_Empleado" in df.columns:
            mias = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
            for _, r in mias.iterrows():
                estado = str(r.get("Estado", "")).strip()
                dias = self._dias_de_solicitud(r)
                if estado == EST_APROBADO:
                    aprobados += dias
                elif estado in (EST_PEND_JEFE, EST_PEND_RRHH, "Pendiente"):
                    en_tramite += dias
        return {"inicial": inicial, "aprobados": aprobados, "en_tramite": en_tramite,
                "usados_total": inicial + aprobados + en_tramite}

    def saldo_vacaciones(self, id_empleado: str, config: dict) -> dict:
        """Resumen completo del derecho a vacaciones de un empleado."""
        emp = self.get_empleado(id_empleado) or {}
        ingreso = self._a_fecha(emp.get("Fecha_Ingreso"))
        usados = self.dias_vacaciones_usados(id_empleado)
        base = {"nombre": str(emp.get("Nombre", id_empleado)),
                "ingreso": ingreso, **usados}
        if not ingreso:
            return {**base, "sin_fecha_ingreso": True, "anios": 0.0,
                    "acumulados": 0.0, "disponibles": 0.0, "derecho_anio_actual": 0}
        anios = self.anios_servicio(ingreso)
        acumulados = self.dias_acumulados(ingreso, config)
        return {**base, "sin_fecha_ingreso": False, "anios": round(anios, 2),
                "acumulados": acumulados,
                "disponibles": round(acumulados - usados["usados_total"], 1),
                "derecho_anio_actual": self.dias_derecho_del_anio(int(anios) + 1, config)}

    def tabla_saldos(self, config: dict) -> pd.DataFrame:
        """Saldo de vacaciones de todos los empleados."""
        df = self.get_empleados()
        if df.empty:
            return pd.DataFrame()
        filas = []
        for _, r in df.iterrows():
            eid = str(r["ID_Empleado"]).strip()
            if not eid:
                continue
            s = self.saldo_vacaciones(eid, config)
            sf = s["sin_fecha_ingreso"]
            # Todo va como texto: mezclar números con "—" en una misma columna
            # rompe la serialización de la tabla en Streamlit.
            filas.append({
                "ID_Empleado": eid,
                "Nombre": s["nombre"],
                "Fecha_Ingreso": s["ingreso"].strftime("%d/%m/%Y") if s["ingreso"] else "—",
                "Años": "—" if sf else f"{s['anios']:.1f}",
                "Derecho_año_actual": "—" if sf else str(s["derecho_anio_actual"]),
                "Acumulados": "—" if sf else f"{s['acumulados']:.1f}",
                "Carga_inicial": f"{s['inicial']:.1f}",
                "Aprobados": f"{s['aprobados']:.1f}",
                "En_trámite": f"{s['en_tramite']:.1f}",
                "Disponibles": "—" if sf else f"{s['disponibles']:.1f}",
            })
        return pd.DataFrame(filas)

    def solicitar_vacaciones(self, id_empleado: str, fecha_ini: str, fecha_fin: str,
                             motivo: str = "", reemplazo: str = "") -> dict:
        """Crea la solicitud de vacaciones. Requiere aprobación del jefe
        inmediato (si está asignado) y luego de RRHH.

        motivo    : razón del pedido, queda registrada en la hoja.
        reemplazo : quién cubre las funciones durante la ausencia.
        """
        dias = self.dias_habiles(fecha_ini, fecha_fin)
        dias_cal = self.dias_calendario(fecha_ini, fecha_fin)
        email_jefe = self.get_email_jefe(id_empleado)
        estado = EST_PEND_JEFE if email_jefe else EST_PEND_RRHH

        vac_id = self._next_id("Vacaciones", "ID_Vacacion", "VAC")
        fila = [vac_id, id_empleado, fecha_ini, fecha_fin, dias, estado,
                "", "", "", "", "", "", "", dias_cal, motivo, reemplazo]
        self.append("Vacaciones", fila)
        return {"id": vac_id, "estado": estado, "dias": dias,
                "dias_calendario": dias_cal, "email_jefe": email_jefe,
                "motivo": motivo, "reemplazo": reemplazo}

    def aprobar_vacaciones_jefe(self, vac_id: str, aprobador: str):
        row = self._fila_de("Vacaciones", "ID_Vacacion", vac_id)
        self.update_campos("Vacaciones", row, {
            "Estado":           EST_PEND_RRHH,
            "Aprobado_Jefe":    aprobador,
            "Fecha_Aprob_Jefe": ahora_local().strftime("%Y-%m-%d %H:%M"),
        })

    def aprobar_vacaciones_rrhh(self, vac_id: str, aprobador: str):
        row = self._fila_de("Vacaciones", "ID_Vacacion", vac_id)
        self.update_campos("Vacaciones", row, {
            "Estado":           EST_APROBADO,
            "Aprobado_RRHH":    aprobador,
            "Fecha_Aprob_RRHH": ahora_local().strftime("%Y-%m-%d %H:%M"),
            "Aprobado_Por":     aprobador,
        })

    def rechazar_vacaciones(self, vac_id: str, rechazado_por: str, motivo: str):
        row = self._fila_de("Vacaciones", "ID_Vacacion", vac_id)
        self.update_campos("Vacaciones", row, {
            "Estado":         EST_RECHAZADO,
            "Motivo_Rechazo": motivo,
            "Rechazado_Por":  rechazado_por,
        })

    # Compatibilidad con la versión anterior
    def aprobar_vacaciones(self, vac_id: str, aprobado_por: str):
        self.aprobar_vacaciones_rrhh(vac_id, aprobado_por)

    def vacaciones_pendientes_jefe(self, email_jefe: str) -> pd.DataFrame:
        ids = self.ids_subordinados(email_jefe)
        df = self.get_vacaciones()
        if df.empty or not ids:
            return pd.DataFrame(columns=HEADERS["Vacaciones"])
        mask = df["ID_Empleado"].astype(str).str.strip().isin(ids) & \
               (df["Estado"].astype(str) == EST_PEND_JEFE)
        return df[mask]

    def vacaciones_pendientes_rrhh(self) -> pd.DataFrame:
        df = self.get_vacaciones()
        if df.empty:
            return pd.DataFrame(columns=HEADERS["Vacaciones"])
        mask = df["Estado"].astype(str).isin([EST_PEND_RRHH, "Pendiente"])
        return df[mask]

    # ── Horas Extras ─────────────────────────────────────────────────────────

    def get_horas_extras(self) -> pd.DataFrame:
        return self.get_df("Horas_Extras")

    def registrar_horas_extra(self, id_empleado: str, fecha: str, horas: float, motivo: str) -> str:
        hex_id = self._next_id("Horas_Extras", "ID", "HE")
        self.append("Horas_Extras", [hex_id, fecha, id_empleado, horas, motivo, "", "Pendiente"])
        return hex_id

    def aprobar_hora_extra(self, hex_id: str, aprobado_por: str):
        row_idx = self._fila_de("Horas_Extras", "ID", hex_id)
        self.update_cell("Horas_Extras", row_idx, 6, aprobado_por)
        self.update_cell("Horas_Extras", row_idx, 7, "Aprobado")

    # ── Usuarios / Autenticación ──────────────────────────────────────────────

    def ensure_usuarios_sheet(self):
        """Crea la hoja Usuarios si no existe.

        Nota de seguridad: ya NO se crea un usuario admin con contraseña fija.
        El admin inicial se crea desde la pantalla de primer arranque, donde la
        persona define su propia contraseña.
        """
        try:
            self._sheet("Usuarios")
        except Exception:
            ws = self.spreadsheet.add_worksheet(title="Usuarios", rows=200, cols=3)
            ws.update(range_name="A1:C1", values=[["ID_Empleado", "Password_Hash", "Rol"]],
                      value_input_option="USER_ENTERED")

    def hay_admin(self) -> bool:
        """True si ya existe al menos un usuario con rol admin y contraseña."""
        df = self.get_usuarios()
        if df.empty or "Rol" not in df.columns:
            return False
        mask = (df["Rol"].astype(str).str.strip().str.lower() == "admin") & \
               (df["Password_Hash"].astype(str).str.strip() != "")
        return bool(mask.any())

    def get_usuarios(self) -> pd.DataFrame:
        return self.get_df("Usuarios")

    def verificar_credenciales(self, id_empleado: str, password: str):
        """Verifica ID + contraseña. Devuelve dict {id_empleado, rol} o None.
        Migra automáticamente los hashes antiguos a PBKDF2 tras un login válido."""
        df = self.get_usuarios()
        if df.empty or "ID_Empleado" not in df.columns:
            return None
        mask = df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()
        rows = df[mask]
        if rows.empty:
            return None
        row = rows.iloc[0]
        almacenado = str(row["Password_Hash"]).strip()
        if not verify_password(password, almacenado):
            return None
        # Migración transparente del hash heredado
        if es_hash_heredado(almacenado):
            try:
                row_idx = self._fila_de("Usuarios", "ID_Empleado", id_empleado)
                self.update_cell("Usuarios", row_idx, 2, hash_password(password))
            except Exception:
                pass
        return {
            "id_empleado": str(id_empleado).strip(),
            "rol": str(row.get("Rol", "empleado")).strip() or "empleado",
        }

    def crear_usuario(self, id_empleado: str, password: str, rol: str = "empleado"):
        pwd_hash = hash_password(password)
        df = self.get_usuarios()
        m = (df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
             if not df.empty and "ID_Empleado" in df.columns else df.iloc[0:0])
        if not m.empty:
            self.update_row("Usuarios", fila_de_hoja(df, m.index[0]),
                            [id_empleado, pwd_hash, rol])
        else:
            self.append("Usuarios", [id_empleado, pwd_hash, rol])

    def cambiar_password(self, id_empleado: str, nueva_password: str):
        row_idx = self._fila_de("Usuarios", "ID_Empleado", id_empleado)
        self.update_cell("Usuarios", row_idx, 2, hash_password(nueva_password))


    # ══════════════════════════════════════════════════════════════════════
    # KYE – Conozca a su Empleado
    # ══════════════════════════════════════════════════════════════════════

    def ensure_hojas_riesgo(self):
        """Crea las hojas de KYE, buró y reconocimientos si no existen."""
        for hoja in ("KYE_Empleado", "Score_Buro", "Reconocimientos",
                     "Documentos", "Capacitaciones", "Titulos"):
            try:
                self._sheet(hoja)
                self.ensure_columns(hoja)
            except Exception:
                try:
                    cols = HEADERS[hoja]
                    ws = self.spreadsheet.add_worksheet(title=hoja, rows=500,
                                                        cols=max(len(cols), 10))
                    ws.update(range_name=f"A1:{rowcol_to_a1(1, len(cols))}",
                              values=[cols], value_input_option="USER_ENTERED")
                except Exception:
                    pass

    def diagnostico_hojas(self) -> list:
        """Estado de cada hoja: cuántas filas tiene y si se puede leer.

        Sirve para distinguir de un vistazo entre "la hoja está vacía" y "la
        hoja no se pudo leer", que en pantalla se veían igual. Todo sale de
        una sola llamada por lotes para no gastar cuota justo cuando se está
        diagnosticando un problema de cuota.
        """
        salida = []
        try:
            hojas = self._hojas(refrescar=True)
        except Exception as e:
            return [{"Hoja": n, "Filas": 0, "Estado": "❌ No se pudo consultar",
                     "Detalle": f"{type(e).__name__}: {e}"} for n in HEADERS]

        presentes = [n for n in HEADERS if n in hojas]
        crudos = {}
        try:
            resp = con_reintentos(self.spreadsheet.values_batch_get,
                                  [f"'{n}'" for n in presentes])
            for nombre, rango in zip(presentes, resp.get("valueRanges", [])):
                crudos[nombre] = rango.get("values", []) or []
        except Exception as e:
            return [{"Hoja": n, "Filas": 0, "Estado": "❌ No se pudo leer",
                     "Detalle": f"{type(e).__name__}: {e}"} for n in HEADERS]

        for nombre in HEADERS:
            fila = {"Hoja": nombre, "Filas": 0, "Estado": "", "Detalle": ""}
            if nombre not in hojas:
                fila["Estado"] = "❌ No existe la pestaña"
                fila["Detalle"] = "Se creará sola al reiniciar la aplicación."
                salida.append(fila)
                continue
            valores = crudos.get(nombre, [])
            cabecera = [str(c).strip() for c in (valores[0] if valores else [])]
            datos = [f for f in valores[1:] if any(str(c).strip() for c in f)]
            fila["Filas"] = len(datos)

            problemas, vistos = [], set()
            for c in cabecera:
                if c and c in vistos:
                    problemas.append(f"encabezado repetido: {c}")
                vistos.add(c)
            vacias = sum(1 for c in cabecera if not c)
            if vacias:
                problemas.append(f"{vacias} encabezado(s) en blanco")
            faltan = [c for c in HEADERS[nombre] if c not in cabecera]
            if faltan:
                problemas.append("faltan columnas: " + ", ".join(faltan))
            fila["Estado"] = "⚠️ Se lee, con avisos" if problemas else "✅ Correcta"
            fila["Detalle"] = " · ".join(problemas)
            salida.append(fila)
        return salida

    def get_kye(self, id_empleado: str) -> dict:
        """Ficha KYE del empleado, o dict vacío si aún no se ha llenado."""
        df = self.get_df("KYE_Empleado")
        if df.empty or "ID_Empleado" not in df.columns:
            return {}
        mask = df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()
        filas = df[mask]
        return filas.iloc[-1].to_dict() if not filas.empty else {}

    def comparar_datos_bloqueo(self, id_empleado: str, nuevos: dict) -> dict:
        """Compara los datos de contacto con los que ya estaban registrados.

        Devuelve {"nuevos": [...], "cambiados": [(etiqueta, antes, ahora)],
                  "sin_cambio": [...], "hay_cambios": bool}
        para que el aviso a sistemas diga qué bloquear y qué liberar.
        """
        previo = self.get_kye(id_empleado)
        altas, cambios, iguales = [], [], []
        for campo, etiqueta in CAMPOS_BLOQUEO_BPM:
            antes  = str(previo.get(campo, "") or "").strip()
            ahora  = str(nuevos.get(campo, "") or "").strip()
            if not ahora:
                continue
            if not antes:
                altas.append((etiqueta, ahora))
            elif antes.lower() != ahora.lower():
                cambios.append((etiqueta, antes, ahora))
            else:
                iguales.append((etiqueta, ahora))
        return {"nuevos": altas, "cambiados": cambios, "sin_cambio": iguales,
                "hay_cambios": bool(altas or cambios),
                "es_primera_vez": not bool(previo)}

    def marcar_aviso_tecnologia(self, id_empleado: str, texto: str):
        """Deja constancia de cuándo se avisó a sistemas."""
        df = self.get_df("KYE_Empleado")
        if df.empty or "ID_Empleado" not in df.columns:
            return
        m = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
        if m.empty:
            return
        self.update_campos("KYE_Empleado", fila_de_hoja(df, m.index[-1]),
                           {"Aviso_Tecnologia": texto})

    def guardar_kye(self, id_empleado: str, datos: dict):
        """Crea o actualiza la ficha KYE del empleado."""
        df = self.get_df("KYE_Empleado")
        cols = HEADERS["KYE_Empleado"]         # nunca incluye la columna interna
        fila = [str(datos.get(c, "")) for c in cols]
        fila[cols.index("ID_Empleado")] = str(id_empleado)
        fila[cols.index("Actualizado")] = ahora_local().strftime("%Y-%m-%d %H:%M")
        m = (df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
             if not df.empty and "ID_Empleado" in df.columns else df.iloc[0:0])
        if not m.empty:
            self.update_row("KYE_Empleado", fila_de_hoja(df, m.index[-1]), fila)
        else:
            self.append("KYE_Empleado", fila)

    # ══════════════════════════════════════════════════════════════════════
    # Score de buró
    # ══════════════════════════════════════════════════════════════════════

    def registrar_score_buro(self, id_empleado: str, score, tipo: str,
                             fuente: str, registrado_por: str,
                             observaciones: str = "", fecha=None) -> str:
        sid = self._next_id("Score_Buro", "ID_Score", "BUR")
        f = (self._a_fecha(fecha) or hoy_local()).strftime("%Y-%m-%d")
        self.append("Score_Buro", [sid, id_empleado, f, score, tipo, fuente,
                                   observaciones, registrado_por])
        return sid

    def historial_buro(self, id_empleado: str = None) -> pd.DataFrame:
        df = self.get_df("Score_Buro")
        if df.empty:
            return df
        if id_empleado:
            df = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
        if "Fecha" in df.columns and not df.empty:
            df = df.assign(_f=df["Fecha"].map(self._a_fecha)).sort_values("_f").drop(columns="_f")
        return df

    def score_buro_actual(self, id_empleado: str) -> dict | None:
        """Último score registrado, y el de ingreso, para ver la evolución."""
        h = self.historial_buro(id_empleado)
        if h.empty:
            return None
        ultimo = h.iloc[-1]
        ingreso = h[h["Tipo"].astype(str).str.lower().str.contains("ingreso")]
        return {
            "actual": self._a_numero(ultimo.get("Score"), 0),
            "fecha": self._a_fecha(ultimo.get("Fecha")),
            "tipo": str(ultimo.get("Tipo", "")),
            "ingreso": self._a_numero(ingreso.iloc[0].get("Score"), 0) if not ingreso.empty else None,
            "revisiones": len(h),
        }

    # ══════════════════════════════════════════════════════════════════════
    # Reconocimientos y felicitaciones
    # ══════════════════════════════════════════════════════════════════════

    def registrar_reconocimiento(self, id_empleado: str, nombre: str, tipo: str,
                                 motivo: str, otorgado_por: str, fecha=None) -> str:
        rid = self._next_id("Reconocimientos", "ID_Reconocimiento", "REC")
        f = (self._a_fecha(fecha) or hoy_local()).strftime("%Y-%m-%d")
        self.append("Reconocimientos", [rid, f, id_empleado, nombre, tipo, motivo,
                                        otorgado_por, "Vigente"])
        return rid

    def get_reconocimientos(self, id_empleado: str = None) -> pd.DataFrame:
        df = self.get_df("Reconocimientos")
        if df.empty or not id_empleado:
            return df
        return df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]

    # ══════════════════════════════════════════════════════════════════════
    # Matriz de riesgo operativo por asesor
    # ══════════════════════════════════════════════════════════════════════
    # El puntaje va de 0 (sin riesgo) a 100 (riesgo máximo). Cada factor
    # devuelve un valor 0..1 que se multiplica por su peso configurable. Todo
    # el desglose se devuelve para que la calificación sea auditable: si un
    # asesor la impugna, se puede mostrar exactamente qué sumó cada factor.

    def _nivel_riesgo(self, puntaje: float, config: dict) -> tuple:
        u_medio = self._a_numero(config.get("Riesgo_Umbral_Medio"), 30)
        u_alto  = self._a_numero(config.get("Riesgo_Umbral_Alto"), 55)
        u_crit  = self._a_numero(config.get("Riesgo_Umbral_Critico"), 75)
        if puntaje >= u_crit:
            return "Crítico", "🔴"
        if puntaje >= u_alto:
            return "Alto", "🟠"
        if puntaje >= u_medio:
            return "Medio", "🟡"
        return "Bajo", "🟢"

    def _f_buro(self, id_empleado: str, config: dict) -> tuple:
        """Riesgo por score de buró. Mayor score crediticio = menor riesgo."""
        s = self.score_buro_actual(id_empleado)
        if not s or not s["actual"]:
            return 0.5, "Sin score de buró registrado (se asume riesgo medio)"
        v = s["actual"]
        bueno = self._a_numero(config.get("Buro_Score_Bueno"), 800)
        malo  = self._a_numero(config.get("Buro_Score_Malo"), 400)
        if v >= bueno:
            f, txt = 0.0, f"Score {v:.0f}: bueno (≥{bueno:.0f})"
        elif v <= malo:
            f, txt = 1.0, f"Score {v:.0f}: deficiente (≤{malo:.0f})"
        else:
            f = (bueno - v) / max(1.0, bueno - malo)
            txt = f"Score {v:.0f}: intermedio entre {malo:.0f} y {bueno:.0f}"
        if s["ingreso"] and v < s["ingreso"] - 50:
            f = min(1.0, f + 0.15)
            txt += f" · deterioro frente al ingreso ({s['ingreso']:.0f})"
        return f, txt

    def _f_pep(self, kye: dict) -> tuple:
        propio = str(kye.get("Es_PEP", "")).strip().lower() in ("si", "sí", "true", "1", "x")
        familiar = str(kye.get("Familiar_PEP", "")).strip().lower() in ("si", "sí", "true", "1", "x")
        if propio and familiar:
            return 1.0, "Es PEP y tiene familiar PEP"
        if propio:
            return 1.0, "Declarado como Persona Expuesta Políticamente"
        if familiar:
            return 0.5, "Tiene familiar o vínculo cercano PEP"
        return 0.0, "Sin condición PEP declarada"

    DOCS_KYE = ["Doc_Hoja_Vida", "Doc_Cedula", "Doc_Cedula_Conyuge", "Doc_Papeleta",
                "Doc_Papeleta_Conyuge", "Doc_Ref_Laborales", "Doc_Ref_Personales",
                "Doc_Servicio_Basico", "Doc_Declaracion_Patrimonial"]

    def _f_documentos(self, kye: dict) -> tuple:
        if not kye:
            return 1.0, "Sin ficha KYE registrada"
        entregados = sum(1 for d in self.DOCS_KYE
                         if str(kye.get(d, "")).strip().lower() in ("si", "sí", "true", "1", "x"))
        total = len(self.DOCS_KYE)
        faltan = total - entregados
        return (faltan / total), f"{entregados} de {total} documentos verificados"

    def _f_familiar(self, kye: dict, config: dict) -> tuple:
        """Presión financiera y estabilidad declaradas en el formulario KYE.

        NOTA: incluye estado civil y número de hijos por decisión expresa del
        área de auditoría. Los pesos son configurables y pueden ponerse en cero.
        """
        if not kye:
            return 0.5, "Sin ficha KYE registrada"
        partes, f = [], 0.0

        civil = str(kye.get("Estado_Civil", "")).strip().lower()
        if civil in ("soltero", "soltero(a)", "divorciado", "divorciado(a)",
                     "separado", "separado(a)", "viudo", "viudo(a)"):
            f += 0.25; partes.append(f"estado civil {civil or 'no declarado'}")
        elif civil:
            partes.append(f"estado civil {civil}")

        hijos = int(self._a_numero(kye.get("Num_Hijos"), 0))
        tope = int(self._a_numero(config.get("Riesgo_Hijos_Tope"), 4))
        if hijos:
            f += 0.25 * min(1.0, hijos / max(1, tope))
            partes.append(f"{hijos} hijo(s)")

        conyuge_trabaja = str(kye.get("Conyuge_Trabaja", "")).strip().lower() in (
            "si", "sí", "true", "1", "x")
        tiene_conyuge = bool(str(kye.get("Conyuge_Nombre", "")).strip())
        if tiene_conyuge and not conyuge_trabaja:
            f += 0.35; partes.append("cónyuge sin ingresos propios")
        elif tiene_conyuge:
            partes.append("cónyuge con ingresos")

        if str(kye.get("Otros_Ingresos", "")).strip().lower() in ("si", "sí", "true", "1", "x"):
            f = max(0.0, f - 0.15); partes.append("declara otros ingresos")

        return min(1.0, f), ("Carga familiar: " + ", ".join(partes) if partes
                             else "Sin datos familiares declarados")

    CAUSALES_CRITICAS = {"Errores de tasación", "Aperturas no autorizadas",
                         "Errores en cierre de caja"}

    def _f_disciplina(self, id_empleado: str, config: dict) -> tuple:
        """Llamados formales y tardanzas de los últimos 12 meses."""
        df = self.llamados_disciplinarios()
        corte = hoy_local(config) - timedelta(days=365)
        puntos, detalle = 0.0, []
        pesos = {"Verbal": 1.0, "Escrito": 3.0, "Suspensión": 6.0}
        if not df.empty and "ID_Empleado" in df.columns:
            mios = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
            for _, r in mios.iterrows():
                f = self._a_fecha(r.get("Fecha"))
                if not f or f < corte:
                    continue
                tipo = str(r.get("Tipo", "")).strip()
                p = pesos.get(tipo, 1.0)
                motivo = str(r.get("Motivo", ""))
                if any(c in motivo for c in self.CAUSALES_CRITICAS):
                    p *= 2
                    detalle.append(f"{tipo} por causal crítica")
                else:
                    detalle.append(tipo)
                puntos += p

        # Los reconocimientos compensan parcialmente
        rec = self.get_reconocimientos(id_empleado)
        n_rec = 0
        if not rec.empty:
            for _, r in rec.iterrows():
                f = self._a_fecha(r.get("Fecha"))
                if f and f >= corte:
                    n_rec += 1
        if n_rec:
            puntos = max(0.0, puntos - n_rec * self._a_numero(
                config.get("Riesgo_Peso_Reconocimiento"), 1.5))
            detalle.append(f"{n_rec} reconocimiento(s) que compensan")

        tope = self._a_numero(config.get("Riesgo_Puntos_Disciplina_Tope"), 12)
        texto = ("Últimos 12 meses: " + ", ".join(detalle)) if detalle else \
                "Sin llamados formales en los últimos 12 meses"
        return min(1.0, puntos / max(1.0, tope)), texto

    def _f_formacion(self, id_empleado: str, config: dict) -> tuple:
        """Formación académica y continuidad educativa.

        Más nivel y capacitación reciente = menor riesgo operativo: alguien
        formado y actualizado comete menos errores de tasación y de proceso.
        """
        p = self.perfil_formacion(id_empleado)
        if p["puntos_nivel"] == 0 and p["cursos_finalizados"] == 0:
            return 1.0, "Sin títulos ni cursos registrados"
        # Nivel académico: 6 puntos (doctorado) reduce el riesgo a cero
        f_nivel = max(0.0, 1 - p["puntos_nivel"] / 6)
        # Formación reciente compensa medio punto
        if p["formacion_reciente"]:
            f_nivel = max(0.0, f_nivel - 0.25)
        partes = [f"nivel: {p['nivel_maximo']}"]
        if p["cursos_finalizados"]:
            partes.append(f"{p['cursos_finalizados']} curso(s) finalizado(s)")
        if p["cursos_en_curso"]:
            partes.append(f"{p['cursos_en_curso']} en curso")
        if p["formacion_reciente"]:
            partes.append("con formación en los últimos 2 años")
        else:
            partes.append("sin formación reciente")
        return f_nivel, ", ".join(partes).capitalize()

    def _f_antiguedad(self, empleado: dict, config: dict) -> tuple:
        ingreso = self._a_fecha(empleado.get("Fecha_Ingreso"))
        if not ingreso:
            return 0.5, "Sin fecha de ingreso registrada"
        anios = self.anios_servicio(ingreso)
        umbral = self._a_numero(config.get("Riesgo_Antiguedad_Anios"), 2)
        if anios >= umbral:
            return 0.0, f"{anios:.1f} años de antigüedad"
        return (1 - anios / max(0.1, umbral)), f"{anios:.1f} años (menos de {umbral:.0f})"

    def evaluar_riesgo(self, id_empleado: str, config: dict) -> dict:
        """Evalúa el riesgo operativo de un asesor.

        Devuelve el puntaje 0-100, el nivel, y el desglose factor por factor
        con el peso, el valor y la explicación de por qué. Todo el cálculo es
        reproducible: mismos datos, mismo resultado.
        """
        empleado = self.get_empleado(id_empleado) or {}
        kye = self.get_kye(id_empleado)

        pesos = {
            "Score de buró":        self._a_numero(config.get("Riesgo_Peso_Buro"), 30),
            "Condición PEP":        self._a_numero(config.get("Riesgo_Peso_PEP"), 10),
            "Documentación KYE":    self._a_numero(config.get("Riesgo_Peso_Documentos"), 15),
            "Situación familiar":   self._a_numero(config.get("Riesgo_Peso_Familiar"), 10),
            "Historial disciplinario": self._a_numero(config.get("Riesgo_Peso_Disciplina"), 25),
            "Antigüedad":           self._a_numero(config.get("Riesgo_Peso_Antiguedad"), 10),
            "Formación":            self._a_numero(config.get("Riesgo_Peso_Formacion"), 10),
        }
        f_buro, t_buro = self._f_buro(id_empleado, config)
        f_pep,  t_pep  = self._f_pep(kye)
        f_doc,  t_doc  = self._f_documentos(kye)
        f_fam,  t_fam  = self._f_familiar(kye, config)
        f_dis,  t_dis  = self._f_disciplina(id_empleado, config)
        f_ant,  t_ant  = self._f_antiguedad(empleado, config)
        f_for,  t_for  = self._f_formacion(id_empleado, config)

        valores = {
            "Score de buró":           (f_buro, t_buro),
            "Condición PEP":           (f_pep,  t_pep),
            "Documentación KYE":       (f_doc,  t_doc),
            "Situación familiar":      (f_fam,  t_fam),
            "Historial disciplinario": (f_dis,  t_dis),
            "Antigüedad":              (f_ant,  t_ant),
            "Formación":               (f_for,  t_for),
        }
        total_peso = sum(pesos.values()) or 1.0
        desglose, puntaje = [], 0.0
        for nombre, (valor, texto) in valores.items():
            peso = pesos[nombre]
            aporte = valor * peso
            puntaje += aporte
            desglose.append({
                "Factor": nombre,
                "Peso": round(peso, 1),
                "Nivel del factor": f"{valor*100:.0f}%",
                "Aporta al puntaje": round(aporte, 1),
                "Por qué": texto,
            })
        # Normalizar a 100 si los pesos no suman 100
        puntaje = round(puntaje * 100 / total_peso, 1)
        nivel, icono = self._nivel_riesgo(puntaje, config)

        return {
            "id_empleado": str(id_empleado),
            "nombre": str(empleado.get("Nombre", id_empleado)),
            "area": str(empleado.get("Area", "")),
            "puntaje": puntaje,
            "nivel": nivel,
            "icono": icono,
            "desglose": desglose,
            "tiene_kye": bool(kye),
            "peso_total": total_peso,
        }

    def matriz_riesgo(self, config: dict) -> pd.DataFrame:
        """Evaluación de riesgo de todo el personal, ordenada de mayor a menor."""
        df = self.get_empleados()
        if df.empty:
            return pd.DataFrame()
        filas = []
        for _, r in df.iterrows():
            eid = str(r["ID_Empleado"]).strip()
            if not eid:
                continue
            ev = self.evaluar_riesgo(eid, config)
            filas.append({
                "ID_Empleado": eid,
                "Nombre": ev["nombre"],
                "Área": ev["area"],
                "Puntaje": ev["puntaje"],
                "Nivel": f"{ev['icono']} {ev['nivel']}",
                "Ficha KYE": "Sí" if ev["tiene_kye"] else "Falta",
            })
        out = pd.DataFrame(filas)
        return out.sort_values("Puntaje", ascending=False) if not out.empty else out


    # ══════════════════════════════════════════════════════════════════════
    # Repositorio de documentos en Google Drive
    # ══════════════════════════════════════════════════════════════════════
    # Streamlit Cloud no guarda archivos: lo que se sube se pierde al
    # reiniciar. Por eso los documentos van a una carpeta de Drive y en la
    # hoja solo se guarda el enlace.
    #
    # IMPORTANTE: una cuenta de servicio no tiene almacenamiento propio en
    # Drive. La carpeta debe ser de una persona real y estar compartida con
    # la cuenta de servicio como Editor; así el espacio se descuenta de esa
    # persona y la subida funciona.

    def _servicio_drive(self):
        from googleapiclient.discovery import build
        if getattr(self, "_drive", None) is None:
            self._drive = build("drive", "v3", credentials=self.credenciales,
                                cache_discovery=False)
        return self._drive

    def subir_documento_drive(self, archivo, nombre_archivo: str,
                              carpeta_id: str) -> dict:
        """Sube un archivo a la carpeta de Drive y devuelve su id y enlace."""
        from googleapiclient.http import MediaIoBaseUpload
        import io, mimetypes

        if not carpeta_id or not str(carpeta_id).strip():
            raise ValueError(
                "No hay carpeta de Drive configurada. Ve a Configuración y pega "
                "el ID de la carpeta donde se guardarán los documentos.")

        contenido = archivo.read() if hasattr(archivo, "read") else archivo
        if not contenido:
            raise ValueError("El archivo está vacío.")

        tipo = mimetypes.guess_type(nombre_archivo)[0] or "application/octet-stream"
        medio = MediaIoBaseUpload(io.BytesIO(contenido), mimetype=tipo, resumable=False)
        meta = {"name": nombre_archivo, "parents": [str(carpeta_id).strip()]}
        creado = con_reintentos(
            lambda: self._servicio_drive().files().create(
                body=meta, media_body=medio,
                fields="id, webViewLink, size").execute())
        return {"id": creado.get("id", ""),
                "link": creado.get("webViewLink", ""),
                "bytes": int(creado.get("size", 0) or 0)}

    def verificar_carpeta_drive(self, carpeta_id: str) -> dict:
        """Comprueba que la carpeta existe y que se puede escribir en ella."""
        if not carpeta_id or not str(carpeta_id).strip():
            return {"ok": False, "mensaje": "No hay carpeta configurada."}
        try:
            info = con_reintentos(
                lambda: self._servicio_drive().files().get(
                    fileId=str(carpeta_id).strip(),
                    fields="id, name, mimeType, capabilities/canAddChildren").execute())
            if info.get("mimeType") != "application/vnd.google-apps.folder":
                return {"ok": False, "mensaje": "Ese ID no corresponde a una carpeta."}
            if not info.get("capabilities", {}).get("canAddChildren"):
                return {"ok": False, "nombre": info.get("name", ""),
                        "mensaje": "La carpeta existe pero la cuenta de servicio no "
                                   "puede escribir en ella. Compártela como Editor."}
            return {"ok": True, "nombre": info.get("name", ""),
                    "mensaje": "Carpeta accesible y con permiso de escritura."}
        except Exception as e:
            return {"ok": False, "mensaje": f"{type(e).__name__}: {e}"}

    def registrar_documento(self, id_empleado: str, nombre: str, categoria: str,
                            tipo_doc: str, nombre_archivo: str, drive_id: str,
                            drive_link: str, subido_por: str, estado: str = "Pendiente",
                            referencia: str = "", observaciones: str = "") -> str:
        did = self._next_id("Documentos", "ID_Documento", "DOC")
        self.append("Documentos", [
            did, hoy_local().strftime("%Y-%m-%d"), id_empleado, nombre, categoria,
            tipo_doc, nombre_archivo, drive_id, drive_link, subido_por,
            estado, "", "", referencia, observaciones])
        return did

    def get_documentos(self, id_empleado: str = None, categoria: str = None,
                       estado: str = None) -> pd.DataFrame:
        df = self.get_df("Documentos")
        if df.empty:
            return df
        if id_empleado:
            df = df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]
        if categoria:
            df = df[df["Categoria"].astype(str) == categoria]
        if estado:
            df = df[df["Estado"].astype(str) == estado]
        return df

    def revisar_documento(self, id_documento: str, aprobado: bool,
                          revisado_por: str, observaciones: str = ""):
        fila = self._fila_de("Documentos", "ID_Documento", id_documento)
        self.update_campos("Documentos", fila, {
            "Estado": "Aprobado" if aprobado else "Rechazado",
            "Revisado_Por": revisado_por,
            "Fecha_Revision": ahora_local().strftime("%Y-%m-%d %H:%M"),
            "Observaciones": observaciones,
        })

    # ══════════════════════════════════════════════════════════════════════
    # Capacitación y títulos
    # ══════════════════════════════════════════════════════════════════════

    def registrar_capacitacion(self, id_empleado: str, nombre: str, curso: str,
                               institucion: str, financiamiento: str,
                               fecha_inicio: str, fecha_fin: str, estado: str,
                               horas, costo, registrado_por: str,
                               doc_id: str = "", observaciones: str = "") -> str:
        cid = self._next_id("Capacitaciones", "ID_Curso", "CAP")
        self.append("Capacitaciones", [
            cid, id_empleado, nombre, curso, institucion, financiamiento,
            fecha_inicio, fecha_fin, estado, horas, costo, doc_id,
            registrado_por, observaciones])
        return cid

    def actualizar_capacitacion(self, id_curso: str, campos: dict):
        fila = self._fila_de("Capacitaciones", "ID_Curso", id_curso)
        self.update_campos("Capacitaciones", fila, campos)

    def get_capacitaciones(self, id_empleado: str = None) -> pd.DataFrame:
        df = self.get_df("Capacitaciones")
        if df.empty or not id_empleado:
            return df
        return df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]

    def registrar_titulo(self, id_empleado: str, nombre: str, nivel: str,
                         titulo: str, institucion: str, anio, senescyt: str,
                         registrado_por: str, doc_id: str = "",
                         observaciones: str = "") -> str:
        tid = self._next_id("Titulos", "ID_Titulo", "TIT")
        self.append("Titulos", [
            tid, id_empleado, nombre, nivel, titulo, institucion, anio, senescyt,
            doc_id, "Pendiente", "", "", registrado_por, observaciones])
        return tid

    def validar_titulo(self, id_titulo: str, valido: bool, validado_por: str,
                       observaciones: str = ""):
        fila = self._fila_de("Titulos", "ID_Titulo", id_titulo)
        self.update_campos("Titulos", fila, {
            "Estado_Validacion": "Validado" if valido else "Observado",
            "Validado_Por": validado_por,
            "Fecha_Validacion": ahora_local().strftime("%Y-%m-%d %H:%M"),
            "Observaciones": observaciones,
        })

    def get_titulos(self, id_empleado: str = None) -> pd.DataFrame:
        df = self.get_df("Titulos")
        if df.empty or not id_empleado:
            return df
        return df[df["ID_Empleado"].astype(str).str.strip() == str(id_empleado).strip()]

    def perfil_formacion(self, id_empleado: str) -> dict:
        """Resumen de la formación del asesor: nivel alcanzado, cursos y vigencia.

        La 'continuidad educativa' mira si hay formación en los últimos dos
        años, que es lo que interesa para sustentar que el personal se mantiene
        actualizado.
        """
        titulos = self.get_titulos(id_empleado)
        cursos  = self.get_capacitaciones(id_empleado)

        nivel_max, puntos_nivel = "Sin registro", 0
        if not titulos.empty:
            validos = titulos[titulos["Estado_Validacion"].astype(str) != "Observado"]
            for _, t in validos.iterrows():
                n = str(t.get("Nivel", "")).strip()
                p = NIVELES_TITULO.get(n, 0)
                if p > puntos_nivel:
                    puntos_nivel, nivel_max = p, n

        corte = hoy_local() - timedelta(days=730)
        recientes = finalizados = en_curso = 0
        horas_total = 0.0
        if not cursos.empty:
            for _, c in cursos.iterrows():
                estado = str(c.get("Estado", "")).strip()
                if estado == "Finalizado":
                    finalizados += 1
                    horas_total += self._a_numero(c.get("Horas"), 0)
                elif estado == "En curso":
                    en_curso += 1
                f = self._a_fecha(c.get("Fecha_Fin")) or self._a_fecha(c.get("Fecha_Inicio"))
                if f and f >= corte:
                    recientes += 1

        return {
            "nivel_maximo": nivel_max,
            "puntos_nivel": puntos_nivel,
            "titulos": 0 if titulos.empty else len(titulos),
            "titulos_validados": 0 if titulos.empty else int(
                (titulos["Estado_Validacion"].astype(str) == "Validado").sum()),
            "cursos_finalizados": finalizados,
            "cursos_en_curso": en_curso,
            "horas_capacitacion": horas_total,
            "formacion_reciente": recientes > 0,
            "cursos_recientes": recientes,
        }

    # ── Llamados de Atención ─────────────────────────────────────────────────

    def ensure_llamados_sheet(self):
        """Crea la hoja Llamados_Atencion si no existe."""
        try:
            self._sheet("Llamados_Atencion")
        except Exception:
            ws = self.spreadsheet.add_worksheet(title="Llamados_Atencion", rows=500, cols=9)
            ws.update(range_name="A1:I1", values=[HEADERS["Llamados_Atencion"]],
                      value_input_option="USER_ENTERED")

    def get_llamados_atencion(self) -> pd.DataFrame:
        return self.get_df("Llamados_Atencion")

    def get_tardanzas_mes(self, id_empleado: str, año_mes: str) -> int:
        df = self.get_df("Asistencia")
        if df.empty or "ID_Empleado" not in df.columns:
            return 0
        mask = (df["ID_Empleado"].astype(str) == str(id_empleado)) & \
               (df["Fecha"].astype(str).str.startswith(año_mes)) & \
               (df["Estado"].astype(str) == "Tardanza")
        return int(len(df[mask]))

    def registrar_tardanza(self, id_empleado: str, nombre: str, hora: str,
                           minutos: int, atrasos_mes: int) -> str:
        """Deja constancia de un atraso en la hoja de Llamados de Atención.

        Es un registro informativo, no una sanción: el tipo lo distingue de los
        llamados que emite RRHH. Sirve para que el atraso quede documentado y
        contabilizado el día que haya que sustentar un llamado formal.
        """
        return self.registrar_llamado_atencion(
            id_empleado, nombre, TIPO_TARDANZA,
            f"Entrada registrada a las {hora} con {minutos} minuto(s) de atraso.",
            atrasos_mes, "Sistema (automático)")

    def llamados_disciplinarios(self) -> pd.DataFrame:
        """Solo los llamados emitidos por RRHH, sin los registros de tardanza."""
        df = self.get_llamados_atencion()
        if df.empty or "Tipo" not in df.columns:
            return df
        return df[df["Tipo"].astype(str) != TIPO_TARDANZA]

    def registros_tardanza(self) -> pd.DataFrame:
        """Solo los registros automáticos de atraso."""
        df = self.get_llamados_atencion()
        if df.empty or "Tipo" not in df.columns:
            return pd.DataFrame(columns=HEADERS["Llamados_Atencion"])
        return df[df["Tipo"].astype(str) == TIPO_TARDANZA]

    def registrar_llamado_atencion(self, id_empleado: str, nombre: str, tipo: str,
                                    motivo: str, atrasos: int, registrado_por: str) -> str:
        año = hoy_local().year
        df = self.get_llamados_atencion()
        if not df.empty and "ID_Llamado" in df.columns:
            este_año = df[df["ID_Llamado"].astype(str).str.startswith(f"LA-{año}-")]
            nums = pd.to_numeric(
                este_año["ID_Llamado"].astype(str).str.extract(r"-(\d+)$")[0],
                errors="coerce"
            ).dropna()
            next_n = int(nums.max()) + 1 if not nums.empty else 1
        else:
            next_n = 1
        llamado_id = f"LA-{año}-{next_n:04d}"
        fecha = hoy_local().strftime("%Y-%m-%d")
        self.append("Llamados_Atencion",
                    [llamado_id, fecha, id_empleado, nombre, tipo, motivo, atrasos, registrado_por, "Activo"])
        return llamado_id
